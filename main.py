r"""
Professional Native + Scanned PDF Bill of Lading Extractor
=========================================================

Run:
    pip install fastapi uvicorn python-multipart pymupdf pillow pytesseract numpy openai pydantic-settings
    uvicorn bl_extractor_native_pdf_app:app --reload --host 0.0.0.0 --port 8000

Environment (.env):
    AZURE_OPENAI_ENDPOINT=https://YOUR-RESOURCE.openai.azure.com/
    AZURE_OPENAI_API_KEY=YOUR_KEY
    AZURE_OPENAI_API_VERSION=2024-08-01-preview
    AZURE_OPENAI_DEPLOYMENT=gpt-4o-mini
    TESSERACT_LANG=eng
    # Windows only, if needed:
    # TESSERACT_CMD=C:\Program Files\Tesseract-OCR\tesseract.exe
"""

from __future__ import annotations

import io
import json
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import numpy as np
import pytesseract
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
try:
    from openai import AzureOpenAI
except Exception:  # Allows /extract-text and local tests to run without the AzureOpenAI class installed.
    AzureOpenAI = None  # type: ignore
from PIL import Image, ImageOps
from pydantic import BaseModel, Field
from pydantic_settings import BaseSettings


# ============================================================
# Settings
# ============================================================

class Settings(BaseSettings):
    azure_openai_endpoint: str = Field(default="", alias="AZURE_OPENAI_ENDPOINT")
    azure_openai_api_key: str = Field(default="", alias="AZURE_OPENAI_API_KEY")
    azure_openai_api_version: str = Field(default="2024-08-01-preview", alias="AZURE_OPENAI_API_VERSION")
    azure_openai_deployment: str = Field(default="gpt-4o-mini", alias="AZURE_OPENAI_DEPLOYMENT")

    ocr_dpi: int = Field(default=300, alias="OCR_DPI")
    tesseract_lang: str = Field(default="eng", alias="TESSERACT_LANG")
    tesseract_cmd: Optional[str] = Field(default=None, alias="TESSERACT_CMD")

    max_input_chars: int = Field(default=90000, alias="MAX_INPUT_CHARS")
    native_min_chars: int = Field(default=600, alias="NATIVE_MIN_CHARS")
    native_min_field_hits: int = Field(default=5, alias="NATIVE_MIN_FIELD_HITS")

    excel_max_rows_per_sheet: int = Field(default=2500, alias="EXCEL_MAX_ROWS_PER_SHEET")
    excel_max_cols_per_sheet: int = Field(default=80, alias="EXCEL_MAX_COLS_PER_SHEET")
    excel_max_cell_chars: int = Field(default=500, alias="EXCEL_MAX_CELL_CHARS")

    class Config:
        env_file = ".env"
        extra = "ignore"


Settings.model_rebuild()
settings = Settings()

if settings.tesseract_cmd:
    pytesseract.pytesseract.tesseract_cmd = settings.tesseract_cmd
elif os.name == "nt":
    for possible in (
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ):
        if os.path.exists(possible):
            pytesseract.pytesseract.tesseract_cmd = possible
            break


# ============================================================
# FastAPI
# ============================================================

app = FastAPI(
    title="Professional PDF + OCR + Excel Bill of Lading Extractor",
    version="3.0.0",
    description="Extracts native/OCR PDF text or Excel sheet data, then extracts structured B/L JSON.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# Models
# ============================================================

class ContainerItem(BaseModel):
    container_number: Optional[str] = None
    seal_number: Optional[str] = None
    container_type: Optional[str] = None
    packages: Optional[str] = None
    gross_weight_kg: Optional[str] = None
    measurement_cbm: Optional[str] = None


class BLEntity(BaseModel):
    document_type: Optional[str] = "Bill of Lading"
    mesco_masterblno: Optional[str] = None
    mesco_bookingnumber: Optional[str] = None
    mesco_acidnumber: Optional[str] = None

    mesco_shippernamecontactno: Optional[str] = None
    mesco_shipperaddress: Optional[str] = None
    mesco_consigneenamecontactno: Optional[str] = None
    mesco_consigneeaddress: Optional[str] = None
    mesco_notify1: Optional[str] = None
    mesco_notifyaddress: Optional[str] = None

    mesco_vessel: Optional[str] = None
    mesco_voytruckno: Optional[str] = None
    mesco_origin: Optional[str] = None
    mesco_destination: Optional[str] = None

    mesco_cargodescription: Optional[str] = None
    cr401_totalgrossweight: Optional[str] = None
    cr401_totalvolume: Optional[str] = None
    cr401_totalpackages: Optional[str] = None
    mesco_nooforgbls: Optional[str] = None

    mesco_containertype: Optional[str] = None
    mesco_containertype2: Optional[str] = None
    mesco_containertype3: Optional[str] = None
    mesco_handlinginformation: Optional[str] = None
    mesco_freightpayableat: Optional[str] = None
    mesco_ponumber: Optional[str] = None
    mesco_customerreference: Optional[str] = None
    mesco_bltype: Optional[int] = None
    mesco_transporttype: Optional[int] = None
    mesco_loadtype: Optional[int] = None
    mesco_direction: Optional[int] = None
    cr401_totalteus: Optional[str] = None

    mesco_pcfreightterm: Optional[str] = None
    mesco_etdorigin: Optional[str] = None
    mesco_etadestination: Optional[str] = None
    mesco_pickupaddress: Optional[str] = None
    mesco_deliveryaddress: Optional[str] = None
    mesco_transhipmentport: Optional[str] = None
    mesco_importerstaxno: Optional[str] = None
    mesco_foreignsupplierregistrationnumber: Optional[str] = None
    mesco_incoterm: Optional[str] = None
    mesco_telexrelease: Optional[bool] = False
    mesco_imoclass: Optional[str] = None
    mesco_unnumber: Optional[str] = None

    mesco_hscode: Optional[str] = None
    mesco_dateofissue: Optional[str] = None
    mesco_placeofissue: Optional[str] = None
    mesco_shippedonboarddate: Optional[str] = None

    container_number: Optional[str] = None
    seal_number: Optional[str] = None
    containers: List[ContainerItem] = Field(default_factory=list)

    extraction_method: Optional[str] = None
    extraction_quality: Dict[str, Any] = Field(default_factory=dict)
    confidence: Dict[str, Any] = Field(default_factory=dict)
    warnings: List[str] = Field(default_factory=list)

    class Config:
        extra = "allow"


REQUIRED_FIELDS = list(BLEntity.model_fields.keys())


def empty_bl_entity() -> Dict[str, Any]:
    data = {k: None for k in REQUIRED_FIELDS}
    data["document_type"] = "Bill of Lading"
    data["containers"] = []
    data["mesco_telexrelease"] = False
    data["extraction_quality"] = {}
    data["confidence"] = {}
    data["warnings"] = []
    return data


# ============================================================
# Native PDF text extraction
# ============================================================

@dataclass
class NativeExtractionResult:
    text: str
    method: str
    page_count: int
    char_count: int
    field_hits: int
    warnings: List[str]


FIELD_KEYWORDS = (
    "bill of lading", "b/l", "booking", "shipper", "consignee", "notify",
    "vessel", "voyage", "port of loading", "port of discharge",
    "place of receipt", "place of delivery", "container", "seal",
    "gross weight", "measurement", "description", "hs code", "acid",
    "freight", "original", "non-negotiable",
)


def normalize_text(text: str) -> str:
    text = text.replace("\x00", " ").replace("\x0c", "\n")
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    return text.strip()


def count_field_hits(text: str) -> int:
    upper = text.lower()
    return sum(1 for kw in FIELD_KEYWORDS if kw in upper)


def _words_to_visual_lines(words: List[Tuple], y_tolerance: float = 3.0) -> List[str]:
    """
    Rebuilds lines from page.get_text("words") using coordinates.
    This is stronger than page.get_text("text") for native PDFs because many B/Ls
    contain columns and absolute-positioned text.
    """
    if not words:
        return []

    # PyMuPDF word tuple:
    # x0, y0, x1, y1, word, block_no, line_no, word_no
    words = sorted(words, key=lambda w: (round(float(w[1]) / y_tolerance), float(w[0])))

    lines: List[List[Tuple]] = []
    current: List[Tuple] = []
    current_y: Optional[float] = None

    for w in words:
        y = float(w[1])
        if current_y is None or abs(y - current_y) <= y_tolerance:
            current.append(w)
            current_y = y if current_y is None else (current_y * 0.75 + y * 0.25)
        else:
            lines.append(sorted(current, key=lambda x: float(x[0])))
            current = [w]
            current_y = y

    if current:
        lines.append(sorted(current, key=lambda x: float(x[0])))

    output: List[str] = []
    for line_words in lines:
        parts: List[str] = []
        prev_x1: Optional[float] = None
        median_width = np.median([max(1.0, float(w[2]) - float(w[0])) for w in line_words]) if line_words else 5.0

        for w in line_words:
            x0, _, x1, _, word = float(w[0]), float(w[1]), float(w[2]), float(w[3]), str(w[4])
            if prev_x1 is not None:
                gap = x0 - prev_x1
                if gap > median_width * 4:
                    parts.append("    ")  # visible column gap
                elif gap > median_width * 1.2:
                    parts.append(" ")
                else:
                    parts.append(" ")
            parts.append(word)
            prev_x1 = x1

        output.append("".join(parts).strip())

    return [ln for ln in output if ln]


def _blocks_to_text(page: fitz.Page) -> str:
    blocks = page.get_text("blocks") or []
    valid = []
    for block in blocks:
        if len(block) >= 5 and isinstance(block[4], str) and block[4].strip():
            block_type = block[6] if len(block) > 6 else 0
            if block_type == 0:
                valid.append(block)
    valid.sort(key=lambda b: (round(float(b[1]), 1), round(float(b[0]), 1)))
    return "\n".join(normalize_text(b[4]) for b in valid if normalize_text(b[4]))


def extract_native_pdf_text(pdf_bytes: bytes) -> NativeExtractionResult:
    """
    Professional native-PDF extraction.
    It returns visual-line text from word coordinates and block text for redundancy.
    """
    warnings: List[str] = []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid PDF: {exc}")

    pages: List[str] = []
    try:
        for i, page in enumerate(doc, start=1):
            words = page.get_text("words") or []
            visual_lines = _words_to_visual_lines(words)
            visual_text = "\n".join(visual_lines)

            block_text = _blocks_to_text(page)

            # Use both. The LLM benefits when field values appear in more than one native ordering.
            candidates = []
            if visual_text.strip():
                candidates.append("[VISUAL WORD ORDER]\n" + visual_text)
            if block_text.strip() and block_text.strip() != visual_text.strip():
                candidates.append("[BLOCK ORDER]\n" + block_text)

            if candidates:
                pages.append(f"--- PAGE {i} ---\n" + "\n\n".join(candidates))
            else:
                warnings.append(f"Page {i}: no native text layer found.")
    finally:
        page_count = len(doc)
        doc.close()

    text = normalize_text("\n\n".join(pages))
    field_hits = count_field_hits(text)
    method = "native" if len(text) >= settings.native_min_chars and field_hits >= settings.native_min_field_hits else "native_sparse"

    if method == "native_sparse":
        warnings.append(
            f"Native text looks sparse: {len(text)} chars and {field_hits} field-keyword hits. OCR/hybrid extraction is recommended."
        )

    return NativeExtractionResult(
        text=text,
        method=method,
        page_count=page_count,
        char_count=len(text),
        field_hits=field_hits,
        warnings=warnings,
    )


# ============================================================
# OCR extraction
# ============================================================

@dataclass
class OcrPageResult:
    page: int
    text: str
    word_count: int
    average_confidence: float


def render_pdf_pages(pdf_bytes: bytes, dpi: int = 300) -> List[Image.Image]:
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid PDF: {exc}")

    pages: List[Image.Image] = []
    zoom = dpi / 72.0
    try:
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            pages.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
    finally:
        doc.close()

    if not pages:
        raise HTTPException(status_code=400, detail="PDF has no readable pages.")
    return pages


def preprocess_for_ocr(image: Image.Image) -> Image.Image:
    gray = ImageOps.grayscale(image)
    gray = ImageOps.autocontrast(gray)
    arr = np.asarray(gray)
    threshold = max(150, int(arr.mean() * 0.92))
    arr = np.where(arr > threshold, 255, arr).astype(np.uint8)
    return Image.fromarray(arr)


def _ocr_text(image: Image.Image, psm: int) -> str:
    return pytesseract.image_to_string(
        image,
        lang=settings.tesseract_lang,
        config=f"--oem 3 --psm {psm}",
    )


def _ocr_data_confidence(image: Image.Image, psm: int) -> Tuple[int, float]:
    data = pytesseract.image_to_data(
        image,
        lang=settings.tesseract_lang,
        config=f"--oem 3 --psm {psm}",
        output_type=pytesseract.Output.DICT,
    )
    confs = []
    word_count = 0
    for txt, conf in zip(data.get("text", []), data.get("conf", [])):
        if not str(txt).strip():
            continue
        try:
            c = float(conf)
        except Exception:
            c = -1
        if c >= 25:
            word_count += 1
        if c >= 0:
            confs.append(c)
    avg = round(sum(confs) / len(confs), 2) if confs else 0.0
    return word_count, avg


def ocr_page(image: Image.Image, page_no: int) -> OcrPageResult:
    processed = preprocess_for_ocr(image)

    # Header and body are handled differently because B/L numbers often live in small top-right boxes.
    width, height = processed.size
    header = processed.crop((0, 0, width, int(height * 0.28)))
    body = processed.crop((0, int(height * 0.22), width, height))

    try:
        header_text = _ocr_text(header, psm=6)
        body_text = _ocr_text(body, psm=4)
        full_text_4 = _ocr_text(processed, psm=4)
        full_text_6 = _ocr_text(processed, psm=6)
        wc4, conf4 = _ocr_data_confidence(processed, psm=4)
        wc6, conf6 = _ocr_data_confidence(processed, psm=6)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Tesseract OCR failed: {exc}. Current tesseract_cmd={pytesseract.pytesseract.tesseract_cmd!r}",
        )

    best_full_text = full_text_4 if conf4 >= conf6 else full_text_6
    word_count = max(wc4, wc6)
    avg_conf = max(conf4, conf6)

    merged = "\n".join(
        [
            "[OCR HEADER PSM6]",
            header_text,
            "[OCR BODY PSM4]",
            body_text,
            "[OCR FULL PAGE BEST]",
            best_full_text,
        ]
    )

    return OcrPageResult(
        page=page_no,
        text=normalize_text(merged),
        word_count=word_count,
        average_confidence=avg_conf,
    )


def ocr_pdf(pdf_bytes: bytes) -> Dict[str, Any]:
    images = render_pdf_pages(pdf_bytes, dpi=settings.ocr_dpi)
    results = [ocr_page(img, i) for i, img in enumerate(images, start=1)]
    merged = "\n\n".join(f"--- PAGE {r.page} ---\n{r.text}" for r in results)
    avg = round(sum(r.average_confidence for r in results) / len(results), 2) if results else 0.0
    return {
        "raw_text": normalize_text(merged),
        "pages": [r.__dict__ for r in results],
        "ocr_summary": {
            "page_count": len(results),
            "average_confidence": avg,
            "total_word_count": sum(r.word_count for r in results),
        },
    }


def extract_pdf_text_professionally(pdf_bytes: bytes, force_ocr: bool = False) -> Dict[str, Any]:
    """
    Best practice:
    1) Try native extraction.
    2) If native is complete, use native.
    3) If native is sparse or force_ocr=True, OCR the PDF.
    4) If both exist, send a HYBRID input to the LLM, with native first and OCR second.
    """
    native = extract_native_pdf_text(pdf_bytes)

    if not force_ocr and native.method == "native":
        return {
            "method": "native",
            "text": native.text,
            "quality": {
                "native_char_count": native.char_count,
                "native_field_hits": native.field_hits,
                "page_count": native.page_count,
                "warnings": native.warnings,
            },
        }

    ocr = ocr_pdf(pdf_bytes)
    ocr_text = ocr["raw_text"]

    if native.text.strip():
        hybrid_text = normalize_text(
            "[NATIVE PDF TEXT]\n"
            + native.text
            + "\n\n[OCR TEXT FALLBACK / VISUAL TEXT]\n"
            + ocr_text
        )
        method = "hybrid_native_ocr"
    else:
        hybrid_text = ocr_text
        method = "ocr"

    return {
        "method": method,
        "text": hybrid_text,
        "quality": {
            "native_char_count": native.char_count,
            "native_field_hits": native.field_hits,
            "page_count": native.page_count,
            "native_warnings": native.warnings,
            "ocr_summary": ocr.get("ocr_summary", {}),
        },
    }


# ============================================================
# Spreadsheet extraction (.xlsx / .xls / .csv)
# ============================================================

SUPPORTED_SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}
SUPPORTED_PDF_EXTENSIONS = {".pdf"}
SUPPORTED_DOCUMENT_EXTENSIONS = SUPPORTED_PDF_EXTENSIONS | SUPPORTED_SPREADSHEET_EXTENSIONS


def _safe_filename(filename: Optional[str]) -> str:
    return Path(filename or "uploaded_file").name


def _file_extension(filename: Optional[str]) -> str:
    return Path(filename or "").suffix.lower()


def normalize_spreadsheet_cell(value: Any) -> str:
    """Convert Excel cell values into compact, AI-friendly strings."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.time() else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return (f"{value:.10f}").rstrip("0").rstrip(".")
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > settings.excel_max_cell_chars:
        text = text[: settings.excel_max_cell_chars].rstrip() + "…"
    return text


def _is_meaningful_spreadsheet_row(values: List[str]) -> bool:
    return any(v.strip() for v in values)


def _row_has_business_terms(values: List[str]) -> bool:
    row_text = " ".join(values).upper()
    terms = (
        "B/L", "BL", "BILL OF LADING", "BOOKING", "MBL", "HBL", "HOUSE", "MASTER",
        "CONTAINER", "SEAL", "SHIPPER", "CONSIGNEE", "NOTIFY", "VESSEL", "VOYAGE",
        "POL", "POD", "PORT", "ORIGIN", "DESTINATION", "GROSS", "WEIGHT", "CBM",
        "MEAS", "PACKAGE", "PCS", "CTNS", "HS", "ACID", "FREIGHT", "PREPAID", "COLLECT",
        "INVOICE", "MANIFEST", "ETA", "ETD", "CFS", "LCL", "FCL",
    )
    return any(term in row_text for term in terms)


def _column_name(index_1_based: int) -> str:
    name = ""
    n = index_1_based
    while n:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name



def _xlsx_col_to_index(col: str) -> int:
    n = 0
    for ch in col.upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
    return n


def _xlsx_cell_ref_to_row_col(ref: str) -> Tuple[int, int]:
    m = re.match(r"([A-Z]+)(\d+)", ref.upper())
    if not m:
        return 1, 1
    return int(m.group(2)), _xlsx_col_to_index(m.group(1))


def _xml_text(element: Optional[ET.Element]) -> str:
    if element is None:
        return ""
    return "".join(element.itertext())


def extract_xlsx_text_zipxml(excel_bytes: bytes, filename: str, openpyxl_error: Optional[Exception] = None) -> Dict[str, Any]:
    """
    Fallback XLSX reader that parses workbook XML directly. This avoids failures caused by
    invalid Excel table definitions that can make openpyxl reject otherwise readable files.
    It reads cell text, shared strings, inline strings, and formulas/results.
    """
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    try:
        zf = zipfile.ZipFile(io.BytesIO(excel_bytes))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid XLSX zip container: {exc}")

    with zf:
        names = set(zf.namelist())
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in names:
            try:
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for si in root.findall(f"{ns_main}si"):
                    shared_strings.append(_xml_text(si))
            except Exception:
                shared_strings = []

        # Map relationship IDs to worksheet XML paths.
        rel_map: Dict[str, str] = {}
        if "xl/_rels/workbook.xml.rels" in names:
            rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            for rel in rel_root.findall(f"{rel_ns}Relationship"):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target", "")
                if rid and "worksheet" in target:
                    if not target.startswith("xl/"):
                        target = "xl/" + target.lstrip("/")
                    rel_map[rid] = target

        sheets: List[Tuple[str, str]] = []
        if "xl/workbook.xml" in names:
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            sheets_node = wb_root.find(f"{ns_main}sheets")
            if sheets_node is not None:
                for sheet in sheets_node.findall(f"{ns_main}sheet"):
                    sheet_name = sheet.attrib.get("name", "Sheet")
                    rid = sheet.attrib.get(f"{ns_rel}id")
                    path = rel_map.get(rid or "")
                    if path and path in names:
                        sheets.append((sheet_name, path))

        if not sheets:
            # Very defensive fallback if workbook relationships are damaged.
            sheet_paths = sorted(n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
            sheets = [(f"Sheet{i}", path) for i, path in enumerate(sheet_paths, start=1)]

        workbook_parts: List[str] = [
            "[SPREADSHEET TEXT]",
            f"Filename: {filename}",
            f"Workbook sheets: {', '.join(name for name, _ in sheets)}",
            "Reader: XLSX XML fallback",
        ]
        if openpyxl_error is not None:
            workbook_parts.append(f"Openpyxl warning: {str(openpyxl_error)[:400]}")

        sheet_summaries: List[Dict[str, Any]] = []
        total_pairs = 0
        all_records: List[Dict[str, Any]] = []

        for sheet_name, sheet_path in sheets:
            try:
                root = ET.fromstring(zf.read(sheet_path))
            except Exception as exc:
                workbook_parts.append(f"--- SHEET: {sheet_name} ---\n[Could not parse sheet XML: {exc}]")
                continue

            rows_map: Dict[int, Dict[int, str]] = {}
            formulas: List[str] = []
            sheet_data = root.find(f"{ns_main}sheetData")
            if sheet_data is not None:
                for row in sheet_data.findall(f"{ns_main}row"):
                    for cell in row.findall(f"{ns_main}c"):
                        ref = cell.attrib.get("r", "A1")
                        r_idx, c_idx = _xlsx_cell_ref_to_row_col(ref)
                        if r_idx > settings.excel_max_rows_per_sheet or c_idx > settings.excel_max_cols_per_sheet:
                            continue
                        cell_type = cell.attrib.get("t")
                        v_node = cell.find(f"{ns_main}v")
                        f_node = cell.find(f"{ns_main}f")
                        is_node = cell.find(f"{ns_main}is")
                        value = ""
                        if cell_type == "s" and v_node is not None:
                            try:
                                idx = int(v_node.text or "0")
                                value = shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
                            except Exception:
                                value = v_node.text or ""
                        elif cell_type == "inlineStr":
                            value = _xml_text(is_node)
                        elif cell_type == "b" and v_node is not None:
                            value = "TRUE" if (v_node.text or "") == "1" else "FALSE"
                        elif v_node is not None:
                            value = v_node.text or ""
                        elif f_node is not None:
                            value = "=" + (f_node.text or "")
                        value = normalize_spreadsheet_cell(value)
                        if value:
                            rows_map.setdefault(r_idx, {})[c_idx] = value
                        if f_node is not None and v_node is not None and len(formulas) < 80:
                            formulas.append(f"{ref}: ={f_node.text or ''} => {normalize_spreadsheet_cell(v_node.text)}")

            max_row = min(max(rows_map.keys(), default=1), settings.excel_max_rows_per_sheet)
            max_col = min(max((max(cols.keys()) for cols in rows_map.values()), default=1), settings.excel_max_cols_per_sheet)
            grid: List[List[str]] = []
            for r in range(1, max_row + 1):
                row_map = rows_map.get(r, {})
                grid.append([row_map.get(c, "") for c in range(1, max_col + 1)])

            records = _extract_table_records_from_grid(sheet_name, grid)
            all_records.extend(records)
            pairs = _detect_key_value_pairs_from_grid(grid, sheet_name)
            total_pairs += len(pairs)
            workbook_parts.append(_grid_to_ai_text(sheet_name, grid, max_col))
            if records:
                workbook_parts.append("[DETECTED SPREADSHEET RECORDS]\n" + "\n".join(
                    f"Record {r['record_index']} from {r['sheet_name']} row {r['source_row']}: "
                    + " | ".join(f"{k}={v}" for k, v in list(r['values_by_header'].items())[:8])
                    for r in records[:50]
                ))
            if pairs:
                workbook_parts.append("[DETECTED KEY/VALUE PAIRS]\n" + "\n".join(pairs[:120]))
            if formulas:
                workbook_parts.append("[FORMULA RESULTS]\n" + "\n".join(formulas[:80]))
            sheet_summaries.append({
                "sheet_name": sheet_name,
                "max_row_seen": max_row,
                "max_col_seen": max_col,
                "non_empty_rows": sum(1 for row in grid if _is_meaningful_spreadsheet_row(row)),
                "detected_key_value_pairs": len(pairs),
                "reader": "xlsx_xml_fallback",
            })

        text = normalize_text("\n\n".join(workbook_parts))
        return {
            "method": "excel_xlsx_xml_fallback",
            "text": text,
            "quality": {
                "document_type_detected": "spreadsheet",
                "sheet_count": len(sheets),
                "sheets": sheet_summaries,
                "char_count": len(text),
                "field_hits": count_field_hits(text),
                "detected_key_value_pairs": total_pairs,
                "detected_record_count": len(all_records),
                "openpyxl_fallback_reason": str(openpyxl_error)[:500] if openpyxl_error is not None else None,
            },
            "records": all_records,
        }


def _detect_key_value_pairs_from_grid(grid: List[List[str]], sheet_name: str, max_pairs: int = 180) -> List[str]:
    """
    Pull obvious key/value pairs from spreadsheet layouts. This helps the LLM with forms,
    invoices, and manifests where labels are not always in a database table format.
    """
    pairs: List[str] = []
    label_regex = re.compile(
        r"\b(B/L|BL|BILL OF LADING|BOOKING|MBL|HBL|SHIPPER|CONSIGNEE|NOTIFY|VESSEL|VOYAGE|"
        r"POL|POD|PORT|ORIGIN|DESTINATION|CONTAINER|SEAL|GROSS|WEIGHT|CBM|MEAS|PACKAGE|"
        r"ACID|HS|FREIGHT|ETA|ETD|INVOICE|REF|REFERENCE|TERMS)\b",
        re.I,
    )
    for r_idx, row in enumerate(grid):
        if len(pairs) >= max_pairs:
            break
        for c_idx, cell in enumerate(row):
            if not cell or not label_regex.search(cell):
                continue
            candidates: List[str] = []
            # same row, next cells
            for k in range(c_idx + 1, min(len(row), c_idx + 5)):
                if row[k]:
                    candidates.append(row[k])
            # same column, cells below
            for rr in range(r_idx + 1, min(len(grid), r_idx + 5)):
                if c_idx < len(grid[rr]) and grid[rr][c_idx]:
                    candidates.append(grid[rr][c_idx])
            if candidates:
                address = f"{_column_name(c_idx + 1)}{r_idx + 1}"
                pairs.append(f"{sheet_name}!{address}: {cell} => {' | '.join(candidates[:4])}")
    return pairs


def _grid_to_ai_text(sheet_name: str, grid: List[List[str]], max_cols: int) -> str:
    """Render a sheet grid as compact TSV plus cell addresses for reliable AI extraction."""
    lines: List[str] = []
    non_empty_rows = 0
    business_rows = 0
    for row_no, row in enumerate(grid, start=1):
        trimmed = row[:max_cols]
        if not _is_meaningful_spreadsheet_row(trimmed):
            continue
        non_empty_rows += 1
        if _row_has_business_terms(trimmed):
            business_rows += 1
        # Keep both human table order and coordinates.
        row_values = "\t".join(trimmed).rstrip()
        addressed = " | ".join(
            f"{_column_name(i + 1)}{row_no}={v}"
            for i, v in enumerate(trimmed)
            if v
        )
        lines.append(f"ROW {row_no}: {row_values}")
        if addressed:
            lines.append(f"CELLS {row_no}: {addressed}")
    if not lines:
        lines.append("[empty sheet]")
    return "\n".join([
        f"--- SHEET: {sheet_name} ---",
        f"Non-empty rows included: {non_empty_rows}; business-keyword rows: {business_rows}",
        *lines,
    ])



# ------------------------------------------------------------
# Spreadsheet multi-record detection
# ------------------------------------------------------------

SPREADSHEET_RECORD_HEADER_TERMS = (
    "H/BL", "HBL", "HOUSE", "B/L", "SHIPPER", "CONSIGNEE", "PLACE OF RECEIPT",
    "PLACE OF DELIVERY", "POL", "POD", "PACKAGE", "PACKING", "GROSS", "WEIGHT",
    "MEAS", "CBM", "CARGO", "DESCRIPTION", "ACID", "HS", "REMARK",
    "STATUS", "TERM", "DELIVERY TERM", "NOMINATED", "FREE HAND", "GATE",
    "NOS. OF", "NOS OF", "MEASUREM", "CARGO TYPE", "H/BL ACID", "RATE",
)


def _normalize_header_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (value or "").upper())


def _spreadsheet_header_score(row: List[str]) -> int:
    score = 0
    seen = set()
    for cell in row:
        up = (cell or "").upper()
        if not up:
            continue
        for term in SPREADSHEET_RECORD_HEADER_TERMS:
            if term in up and term not in seen:
                score += 1
                seen.add(term)
    return score


def _looks_like_total_or_summary_row(row: List[str]) -> bool:
    row_text = " ".join(v for v in row if v).strip().upper()
    if not row_text:
        return True
    # Check if row starts with TOTAL (handles TOTAL:, TOTAL CIF CARGO, TOTAL FOB CARGO, etc.)
    if re.search(r"^\s*TOTAL\b", row_text):
        return True
    # Check first non-empty cell for TOTAL/SUBTOTAL/GRAND TOTAL (with optional colon)
    first = next((v.strip().upper() for v in row if v.strip()), "")
    if re.match(r"^TOTAL\s*:?$|^SUBTOTAL\s*:?$|^GRAND\s+TOTAL\s*:?$", first):
        return True
    if first in {"TOTAL", "SUBTOTAL", "GRAND TOTAL"}:
        return True
    # Match pattern like "TOTAL 1234.56" or "TOTAL: 1234.56"
    if re.fullmatch(r"TOTAL\s*:?(?:\s+[0-9.,]+)*", row_text):
        return True
    return False


def _global_context_from_grid(sheet_name: str, grid: List[List[str]], header_row_index_0: int, max_lines: int = 12) -> List[str]:
    """Rows above the record header usually contain schedule/global shipment context."""
    lines: List[str] = []
    for row_no, row in enumerate(grid[:header_row_index_0], start=1):
        if not _is_meaningful_spreadsheet_row(row):
            continue
        row_values = "\t".join(row).rstrip()
        addressed = " | ".join(f"{_column_name(i + 1)}{row_no}={v}" for i, v in enumerate(row) if v)
        if row_values:
            lines.append(f"{sheet_name} ROW {row_no}: {row_values}")
        if addressed:
            lines.append(f"{sheet_name} CELLS {row_no}: {addressed}")
        if len(lines) >= max_lines:
            break
    return lines


def _is_data_row_not_header(row: List[str]) -> bool:
    """
    Returns True if this row looks like a data row rather than a header.
    Data rows have long free-text cells (addresses, cargo descriptions),
    whereas header rows have short label-like cells.
    """
    long_cells = sum(1 for v in row if len(v) > 40)
    return long_cells >= 2


def _extract_table_records_from_grid(sheet_name: str, grid: List[List[str]], max_records: int = 300) -> List[Dict[str, Any]]:
    """
    Detect manifest/loading-sheet style tables and split every shipment row into its own record.
    Supports MULTIPLE header sections (e.g., CIF section with header, then FOB section with header).

    Example supported layout:
      header row: H/BL Nos. | Term | Delivery Term | Status | Shipper | Consignee | ...
      data rows: one HBL/shipment per row
      [optional blank rows]
      second header row: H/BL Nos. | Term | ...  (e.g., FOB section)
      more data rows

    Returns record dicts with a dedicated AI text block so Azure extracts one CRM payload per row,
    instead of mixing multiple HBLs into one response.
    """
    if not grid:
        return []

    # Find ALL header rows with score >= 3 (lowered from 5 → 3 due to Chinese filler columns)
    header_candidates = []
    for idx, row in enumerate(grid):
        score = _spreadsheet_header_score(row)
        if score >= 3:
            headers = [h.strip() for h in row]
            # Exclude Chinese filler columns (列1, 列2, etc.)
            non_empty_header_count = sum(1 for h in headers if h and not re.match(r'^列\d+$', h))
            # CRITICAL: reject rows that look like data rows (long free-text cells)
            if non_empty_header_count >= 4 and not _is_data_row_not_header(row):
                header_candidates.append((idx, score))

    if not header_candidates:
        return []

    records: List[Dict[str, Any]] = []

    # Process each header section independently
    for section_num, (best_idx, _score) in enumerate(header_candidates):
        headers = [h.strip() for h in grid[best_idx]]
        global_context = _global_context_from_grid(sheet_name, grid, best_idx)

        # Determine data range: from after this header until next header (or end of grid)
        data_start = best_idx + 1
        data_end = len(grid)
        if section_num + 1 < len(header_candidates):
            data_end = header_candidates[section_num + 1][0]  # Stop before next header

        for row_idx in range(data_start, data_end):
            row = grid[row_idx]
            if not _is_meaningful_spreadsheet_row(row):
                # Do not stop immediately; some sheets have spacer rows inside tables.
                continue
            if _looks_like_total_or_summary_row(row):
                continue

            values_by_header: Dict[str, str] = {}
            filled_cells = 0
            for col_idx, header in enumerate(headers):
                if not header or col_idx >= len(row):
                    continue
                value = row[col_idx].strip()
                if not value:
                    continue
                values_by_header[header] = value
                filled_cells += 1

            row_text_upper = " ".join(row).upper()
            has_record_identifier = bool(
                re.search(r"\b[A-Z]{2,}[A-Z0-9]*\d{4,}[A-Z0-9]*\b", row_text_upper)
                or re.search(r"\b\d{10,19}\b", row_text_upper)
                or any(h and v for h, v in values_by_header.items() if re.search(r"H\s*/?\s*BL|HOUSE|B/L|BL\s*NO", h, re.I))
            )
            has_party = any(re.search(r"SHIPPER|CONSIGNEE", h, re.I) for h in values_by_header)

            if filled_cells < 3 or not (has_record_identifier or has_party):
                continue

            record_index = len(records) + 1
            header_line = "\t".join(headers).rstrip()
            row_line = "\t".join(row[:len(headers)]).rstrip()
            mapped_lines = []
            for header, value in values_by_header.items():
                col_idx = headers.index(header)
                mapped_lines.append(f"{_column_name(col_idx + 1)}{row_idx + 1} | {header} => {value}")

            record_text = normalize_text("\n".join([
                "[SPREADSHEET SINGLE RECORD]",
                f"Sheet: {sheet_name}",
                f"Record index: {record_index}",
                f"Source Excel row: {row_idx + 1}",
                "Instruction: Extract ONLY this one spreadsheet row as ONE Mesco CRM shipment/B/L payload. Do not mix values from other data rows.",
                "[GLOBAL CONTEXT FROM ROWS ABOVE TABLE]",
                *(global_context or ["No global context detected."]),
                "[TABLE HEADER]",
                f"ROW {best_idx + 1}: {header_line}",
                "[THIS RECORD ROW]",
                f"ROW {row_idx + 1}: {row_line}",
                "[HEADER TO VALUE MAP FOR THIS RECORD]",
                *mapped_lines,
            ]))

            records.append({
                "record_index": record_index,
                "sheet_name": sheet_name,
                "header_row": best_idx + 1,
                "source_row": row_idx + 1,
                "values_by_header": values_by_header,
                "text": record_text,
            })
            if len(records) >= max_records:
                break
        
        # Stop if we've already reached max records across all sections
        if len(records) >= max_records:
            break

    return records


def extract_xlsx_text(excel_bytes: bytes, filename: str) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Missing dependency for .xlsx files. Install: pip install openpyxl") from exc

    try:
        wb_values = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=False)
        wb_formulas = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False, read_only=False)
    except Exception as exc:
        return extract_xlsx_text_zipxml(excel_bytes, filename, openpyxl_error=exc)

    workbook_parts: List[str] = [
        "[SPREADSHEET TEXT]",
        f"Filename: {filename}",
        f"Workbook sheets: {', '.join(wb_values.sheetnames)}",
        "Instructions for AI: Values are presented with row numbers and cell addresses. Use explicit labels and nearby values.",
    ]
    all_pairs: List[str] = []
    all_records: List[Dict[str, Any]] = []
    sheet_summaries: List[Dict[str, Any]] = []

    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        ws_formula = wb_formulas[sheet_name]
        max_row = min(ws.max_row or 1, settings.excel_max_rows_per_sheet)
        max_col = min(ws.max_column or 1, settings.excel_max_cols_per_sheet)
        grid: List[List[str]] = []
        formula_notes: List[str] = []

        for r in range(1, max_row + 1):
            row_vals: List[str] = []
            for c in range(1, max_col + 1):
                value = normalize_spreadsheet_cell(ws.cell(r, c).value)
                formula_value = ws_formula.cell(r, c).value
                if isinstance(formula_value, str) and formula_value.startswith("=") and value:
                    # Keep formula result only in table, but record the formula once for traceability.
                    if len(formula_notes) < 80:
                        formula_notes.append(f"{_column_name(c)}{r}: {formula_value} => {value}")
                row_vals.append(value)
            grid.append(row_vals)

        records = _extract_table_records_from_grid(sheet_name, grid)
        all_records.extend(records)
        pairs = _detect_key_value_pairs_from_grid(grid, sheet_name)
        all_pairs.extend(pairs)
        workbook_parts.append(_grid_to_ai_text(sheet_name, grid, max_col))
        if records:
            workbook_parts.append("[DETECTED SPREADSHEET RECORDS]\n" + "\n".join(
                f"Record {r['record_index']} from {r['sheet_name']} row {r['source_row']}: "
                + " | ".join(f"{k}={v}" for k, v in list(r['values_by_header'].items())[:8])
                for r in records[:50]
            ))
        if pairs:
            workbook_parts.append("[DETECTED KEY/VALUE PAIRS]\n" + "\n".join(pairs[:120]))
        if formula_notes:
            workbook_parts.append("[FORMULA RESULTS]\n" + "\n".join(formula_notes[:80]))

        sheet_summaries.append({
            "sheet_name": sheet_name,
            "max_row_seen": max_row,
            "max_col_seen": max_col,
            "non_empty_rows": sum(1 for row in grid if _is_meaningful_spreadsheet_row(row)),
            "detected_key_value_pairs": len(pairs),
        })

    text = normalize_text("\n\n".join(workbook_parts))
    return {
        "method": "excel_xlsx",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": len(wb_values.sheetnames),
            "sheets": sheet_summaries,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_key_value_pairs": len(all_pairs),
            "detected_record_count": len(all_records),
        },
        "records": all_records,
    }


def extract_xls_text(excel_bytes: bytes, filename: str) -> Dict[str, Any]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Missing dependency for .xls files. Install: pip install xlrd") from exc

    try:
        book = xlrd.open_workbook(file_contents=excel_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid XLS workbook: {exc}")

    workbook_parts: List[str] = [
        "[SPREADSHEET TEXT]",
        f"Filename: {filename}",
        f"Workbook sheets: {', '.join(book.sheet_names())}",
        "Instructions for AI: Values are presented with row numbers and cell addresses. Use explicit labels and nearby values.",
    ]
    sheet_summaries: List[Dict[str, Any]] = []
    total_pairs = 0
    all_records: List[Dict[str, Any]] = []

    for sheet in book.sheets():
        max_row = min(sheet.nrows, settings.excel_max_rows_per_sheet)
        max_col = min(sheet.ncols, settings.excel_max_cols_per_sheet)
        grid: List[List[str]] = []
        for r in range(max_row):
            row_vals: List[str] = []
            for c in range(max_col):
                cell = sheet.cell(r, c)
                value: Any = cell.value
                # Convert Excel serial dates when xlrd marks them as dates.
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                row_vals.append(normalize_spreadsheet_cell(value))
            grid.append(row_vals)
        records = _extract_table_records_from_grid(sheet.name, grid)
        all_records.extend(records)
        pairs = _detect_key_value_pairs_from_grid(grid, sheet.name)
        total_pairs += len(pairs)
        workbook_parts.append(_grid_to_ai_text(sheet.name, grid, max_col))
        if records:
            workbook_parts.append("[DETECTED SPREADSHEET RECORDS]\n" + "\n".join(
                f"Record {r['record_index']} from {r['sheet_name']} row {r['source_row']}: "
                + " | ".join(f"{k}={v}" for k, v in list(r['values_by_header'].items())[:8])
                for r in records[:50]
            ))
        if pairs:
            workbook_parts.append("[DETECTED KEY/VALUE PAIRS]\n" + "\n".join(pairs[:120]))
        sheet_summaries.append({
            "sheet_name": sheet.name,
            "max_row_seen": max_row,
            "max_col_seen": max_col,
            "non_empty_rows": sum(1 for row in grid if _is_meaningful_spreadsheet_row(row)),
            "detected_key_value_pairs": len(pairs),
        })

    text = normalize_text("\n\n".join(workbook_parts))
    return {
        "method": "excel_xls",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": book.nsheets,
            "sheets": sheet_summaries,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_key_value_pairs": total_pairs,
            "detected_record_count": len(all_records),
        },
        "records": all_records,
    }


def extract_csv_text(csv_bytes: bytes, filename: str) -> Dict[str, Any]:
    import csv
    text_raw = csv_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text_raw))
    rows = []
    for i, row in enumerate(reader, start=1):
        if i > settings.excel_max_rows_per_sheet:
            break
        clean = [normalize_spreadsheet_cell(v) for v in row[: settings.excel_max_cols_per_sheet]]
        if _is_meaningful_spreadsheet_row(clean):
            rows.append(f"ROW {i}: " + "\t".join(clean))
    text = normalize_text("\n".join([
        "[SPREADSHEET TEXT]",
        f"Filename: {filename}",
        "--- SHEET: CSV ---",
        *rows,
    ]))
    return {
        "method": "csv",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": 1,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_record_count": 0,
        },
        "records": [],
    }


def extract_spreadsheet_text_professionally(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    ext = _file_extension(filename)
    if ext == ".xlsx":
        return extract_xlsx_text(file_bytes, filename)
    if ext == ".xls":
        return extract_xls_text(file_bytes, filename)
    if ext == ".csv":
        return extract_csv_text(file_bytes, filename)
    raise HTTPException(status_code=400, detail=f"Unsupported spreadsheet type: {ext}")


def extract_document_text_professionally(file_bytes: bytes, filename: str, force_ocr: bool = False) -> Dict[str, Any]:
    """Route uploaded PDFs and spreadsheets into one common AI text format."""
    ext = _file_extension(filename)
    if ext in SUPPORTED_PDF_EXTENSIONS:
        result = extract_pdf_text_professionally(file_bytes, force_ocr=force_ocr)
        result["source_file_type"] = "pdf"
        return result
    if ext in SUPPORTED_SPREADSHEET_EXTENSIONS:
        result = extract_spreadsheet_text_professionally(file_bytes, filename)
        result["source_file_type"] = "spreadsheet"
        return result
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported file type '{ext}'. Upload PDF, XLSX, XLS, or CSV.",
    )


# ============================================================
# Azure OpenAI structured extraction
# ============================================================

JSON_SCHEMA = {
    "name": "bill_of_lading_extraction",
    "schema": {
        "type": "object",
        "additionalProperties": True,
        "properties": {
            "document_type": {"type": ["string", "null"]},
            "mesco_masterblno": {"type": ["string", "null"]},
            "mesco_bookingnumber": {"type": ["string", "null"]},
            "mesco_acidnumber": {"type": ["string", "null"]},
            "mesco_shippernamecontactno": {"type": ["string", "null"]},
            "mesco_shipperaddress": {"type": ["string", "null"]},
            "mesco_consigneenamecontactno": {"type": ["string", "null"]},
            "mesco_consigneeaddress": {"type": ["string", "null"]},
            "mesco_notify1": {"type": ["string", "null"]},
            "mesco_notifyaddress": {"type": ["string", "null"]},
            "mesco_vessel": {"type": ["string", "null"]},
            "mesco_voytruckno": {"type": ["string", "null"]},
            "mesco_origin": {"type": ["string", "null"]},
            "mesco_destination": {"type": ["string", "null"]},
            "mesco_cargodescription": {"type": ["string", "null"]},
            "cr401_totalgrossweight": {"type": ["string", "null"]},
            "cr401_totalvolume": {"type": ["string", "null"]},
            "cr401_totalpackages": {"type": ["string", "null"]},
            "mesco_nooforgbls": {"type": ["string", "null"]},
            "mesco_containertype": {"type": ["string", "null"]},
            "mesco_containertype2": {"type": ["string", "null"]},
            "mesco_containertype3": {"type": ["string", "null"]},
            "mesco_handlinginformation": {"type": ["string", "null"]},
            "mesco_freightpayableat": {"type": ["string", "null"]},
            "mesco_ponumber": {"type": ["string", "null"]},
            "mesco_customerreference": {"type": ["string", "null"]},
            "mesco_bltype": {"type": ["integer", "null"]},
            "mesco_transporttype": {"type": ["integer", "null"]},
            "mesco_loadtype": {"type": ["integer", "null"]},
            "mesco_direction": {"type": ["integer", "null"]},
            "cr401_totalteus": {"type": ["string", "null"]},
            "mesco_pcfreightterm": {"type": ["string", "null"]},
            "mesco_etdorigin": {"type": ["string", "null"]},
            "mesco_etadestination": {"type": ["string", "null"]},
            "mesco_pickupaddress": {"type": ["string", "null"]},
            "mesco_deliveryaddress": {"type": ["string", "null"]},
            "mesco_transhipmentport": {"type": ["string", "null"]},
            "mesco_importerstaxno": {"type": ["string", "null"]},
            "mesco_foreignsupplierregistrationnumber": {"type": ["string", "null"]},
            "mesco_incoterm": {"type": ["string", "null"]},
            "mesco_telexrelease": {"type": ["boolean", "null"]},
            "mesco_imoclass": {"type": ["string", "null"]},
            "mesco_unnumber": {"type": ["string", "null"]},
            "mesco_hscode": {"type": ["string", "null"]},
            "mesco_dateofissue": {"type": ["string", "null"]},
            "mesco_placeofissue": {"type": ["string", "null"]},
            "mesco_shippedonboarddate": {"type": ["string", "null"]},
            "container_number": {"type": ["string", "null"]},
            "seal_number": {"type": ["string", "null"]},
            "containers": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "container_number": {"type": ["string", "null"]},
                        "seal_number": {"type": ["string", "null"]},
                        "container_type": {"type": ["string", "null"]},
                        "packages": {"type": ["string", "null"]},
                        "gross_weight_kg": {"type": ["string", "null"]},
                        "measurement_cbm": {"type": ["string", "null"]},
                    },
                    "required": [
                        "container_number", "seal_number", "container_type",
                        "packages", "gross_weight_kg", "measurement_cbm"
                    ],
                },
            },
            "confidence": {"type": "object"},
            "warnings": {"type": "array", "items": {"type": "string"}},
        },
        "required": [
            "document_type", "mesco_masterblno", "mesco_bookingnumber", "mesco_acidnumber",
            "mesco_shippernamecontactno", "mesco_shipperaddress",
            "mesco_consigneenamecontactno", "mesco_consigneeaddress",
            "mesco_notify1", "mesco_notifyaddress", "mesco_vessel", "mesco_voytruckno",
            "mesco_origin", "mesco_destination", "mesco_cargodescription",
            "cr401_totalgrossweight", "cr401_totalvolume", "cr401_totalpackages",
            "mesco_nooforgbls", "mesco_containertype", "mesco_containertype2",
            "mesco_containertype3", "mesco_handlinginformation",
            "mesco_freightpayableat", "mesco_ponumber", "mesco_customerreference",
            "mesco_bltype", "mesco_transporttype", "mesco_loadtype", "mesco_direction",
            "cr401_totalteus", "mesco_pcfreightterm", "mesco_etdorigin",
            "mesco_etadestination", "mesco_pickupaddress", "mesco_deliveryaddress",
            "mesco_transhipmentport", "mesco_importerstaxno",
            "mesco_foreignsupplierregistrationnumber", "mesco_incoterm",
            "mesco_telexrelease", "mesco_imoclass", "mesco_unnumber", "mesco_hscode",
            "mesco_dateofissue", "mesco_placeofissue", "mesco_shippedonboarddate",
            "container_number", "seal_number", "containers", "confidence", "warnings",
        ],
    },
}


SYSTEM_PROMPT = """
You are a professional Bill of Lading extraction engine for Mesco CRM / Dynamics 365.

You receive text extracted from a PDF or spreadsheet. It may be:
- native PDF text from coordinate-based extraction,
- OCR text from a scanned/image PDF,
- hybrid text containing both,
- or spreadsheet text from XLSX/XLS/CSV rows, cells, formulas, and detected key/value pairs.

Return ONLY valid JSON that follows the provided schema.

Important source rules:
- Native PDFs often store text out of visual order. Use labels and nearby values, not only line order.
- If there are [NATIVE PDF TEXT] and [OCR TEXT FALLBACK / VISUAL TEXT] sections, prefer the value that is most clearly attached to a field label.
- OCR may reveal visual text that native text misses.
- Spreadsheets are presented as [SPREADSHEET TEXT] with ROW lines, CELLS lines, and DETECTED KEY/VALUE PAIRS. Use sheet names, row numbers, cell addresses, and nearby headers to map values.
- For spreadsheet manifests/invoices, a B/L number may be under HBL, HBL NO, HOUSE B/L, MBL, MASTER B/L, B/L NO, or BL NO. Map the actual B/L number to mesco_masterblno when it is the main shipment B/L. Use booking fields only when explicitly a booking number.

Critical Bill of Lading rules:
- mesco_masterblno is the B/L number. It is ALWAYS found near the label "B/L No.", "B/L NO", "BL NO", or "BILL OF LADING NO".
- CRITICAL DISAMBIGUATION: The shipper company name (e.g., "SWEDEV AB") is NEVER the B/L number. The B/L number appears on the RIGHT side near the carrier logo.
- PURE NUMERIC B/L numbers with spaces (e.g., "85 008") are valid when they appear directly after a B/L label. Preserve the space in the value (e.g., return "85 008").
- Do not confuse B/L number with booking number, ACID number, container number, postal code, company name, or address.
- mesco_bookingnumber is booking/reference number when explicitly labeled booking number.
- mesco_acidnumber is the Egyptian ACID number. Return digits only.
- Extract shipper, consignee, notify party, vessel, voyage, origin/loading port, destination/discharge port.
- Vessel names must not include the port of loading. If text says "CMA CGM TIGA 0NVLOS1MA HAMBURG", vessel is "CMA CGM TIGA", voyage is "0NVLOS1MA", origin is "HAMBURG".
- Extract container numbers as 4 letters + 7 digits, seal numbers, container type, packages, gross weight, and CBM.
- Freight terms: if "FREIGHT COLLECT", set mesco_pcfreightterm="COLLECT"; if "FREIGHT PREPAID", set "PREPAID".
- mesco_transporttype must be 300000000 for sea B/Ls.
- mesco_loadtype: FCL/container present = 300000000; LCL explicitly stated = 300000001.
- mesco_direction: import = 300000000 if destination/discharge is Egypt; export = 300000001 if origin/loading is Egypt.
- mesco_bltype: if ORIGINAL appears and it is not only boilerplate, use 886150001. If NON-NEGOTIABLE is prominent, leave null unless your CRM mapping requires another value.
- Do not guess. Use null when a value is not visible.
"""


def get_azure_client() -> AzureOpenAI:
    if AzureOpenAI is None:
        raise HTTPException(
            status_code=500,
            detail="Azure OpenAI SDK is not available. Install/upgrade with: pip install --upgrade openai",
        )
    if not settings.azure_openai_endpoint or not settings.azure_openai_api_key:
        raise HTTPException(
            status_code=500,
            detail="Azure OpenAI is not configured. Set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY.",
        )
    return AzureOpenAI(
        azure_endpoint=settings.azure_openai_endpoint,
        api_key=settings.azure_openai_api_key,
        api_version=settings.azure_openai_api_version,
    )


def extract_with_azure_openai(extracted_text: str) -> Dict[str, Any]:
    text = normalize_text(extracted_text)
    if not text:
        raise HTTPException(status_code=400, detail="No text was extracted from the PDF.")

    client = get_azure_client()
    safe_text = text[: settings.max_input_chars]

    try:
        response = client.chat.completions.create(
            model=settings.azure_openai_deployment,
            temperature=0,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": "Extract this Bill of Lading:\n\n" + safe_text},
            ],
            response_format={"type": "json_schema", "json_schema": JSON_SCHEMA},
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI extraction failed: {exc}")

    content = response.choices[0].message.content
    if not content:
        raise HTTPException(status_code=502, detail="Azure OpenAI returned empty content.")

    try:
        data = json.loads(content)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI returned invalid JSON: {exc}")

    return validate_and_correct(data, extracted_text)


# ============================================================
# Deterministic post-processing and regex fallback
# ============================================================

EGYPT_PORTS = {
    "ALEXANDRIA", "PORT SAID", "DAMIETTA", "SOKHNA", "EL SOKHNA",
    "DEKHEILA", "ADABIYA", "PORT SAID WEST", "AIN SOKHNA",
}

KNOWN_PORTS = {
    *EGYPT_PORTS,
    "HAMBURG", "ROTTERDAM", "GOTHENBURG", "ANTWERP", "SHANGHAI",
    "SINGAPORE", "JEBEL ALI", "PIRAEUS", "FELIXSTOWE", "LE HAVRE",
    "VALENCIA", "BARCELONA", "GENOA", "ISTANBUL", "MERSIN", "AMBARLI",
    "COLOMBO", "PORT KLANG", "TANJUNG PELEPAS", "BUSAN", "NINGBO",
    "QINGDAO", "TIANJIN", "HONG KONG", "YOKOHAMA", "NHAVA SHEVA",
    "DURBAN", "MOMBASA", "LAGOS", "AQABA", "BEIRUT", "HAIFA",
    "LIMASSOL", "MALTA", "GIOIA TAURO",
}

BL_BLACKLIST = {
    "ORIGINAL", "TELEX", "RELEASE", "EXPRESS", "FREIGHT", "COLLECT", "PREPAID",
    "SHIPPER", "CONSIGNEE", "NOTIFY", "VESSEL", "VOYAGE", "CONTAINER", "SEAL",
    "WEIGHT", "MEASUREMENT", "DESCRIPTION", "GOODS", "PORT", "LOADING",
    "DISCHARGE", "DELIVERY", "RECEIPT", "PLACE", "DATE", "NUMBER", "BILL",
    "LADING", "BOOKING", "REFERENCE", "ACID", "ORDER", "INVOICE", "PACKING",
    "GROSS", "NET", "TARE", "TOTAL", "PAGE", "COPY", "DRAFT", "SIGNED",
}


def add_warning(data: Dict[str, Any], message: str) -> None:
    data.setdefault("warnings", [])
    if message not in data["warnings"]:
        data["warnings"].append(message)


def clean_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    value = re.sub(r"\s+", " ", str(value)).strip(" ,;:-")
    return value or None


def normalize_digits(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    digits = re.sub(r"\D", "", value)
    return digits or None


def is_container_number(value: Optional[str]) -> bool:
    return bool(value and re.fullmatch(r"[A-Z]{4}\d{7}", value.strip().upper().replace(" ", "")))


def normalize_numeric(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    # 335,000 in B/L often means 335.000 kg; keep decimal style.
    if "," in text and "." not in text:
        parts = text.split(",")
        text = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) == 3 else text.replace(",", ".")
    else:
        text = text.replace(",", "")
    m = re.search(r"\d+(?:\.\d+)?", text)
    return m.group(0) if m else None


def is_likely_bl_number(candidate: str, current_acid: Optional[str] = None) -> bool:
    c = candidate.upper().strip()
    c_compact = re.sub(r"[\s\-]", "", c)

    if len(c_compact) < 5 or len(c_compact) > 25:
        return False
    if c in BL_BLACKLIST or c_compact in BL_BLACKLIST:
        return False
    if is_container_number(c_compact):
        return False
    if current_acid and c_compact == re.sub(r"\D", "", current_acid):
        return False
    if c_compact.isdigit() and len(c_compact) == 19:
        return False

    # NEW: Reject if it looks like a company/person name (all letters, no numbers)
    # unless it's a known carrier code format (3-4 letters + numbers)
    if c_compact.isalpha():
        # Pure alphabetic — only accept carrier-code-like patterns (3-4 uppercase letters as prefix)
        if not re.match(r'^[A-Z]{3,4}[A-Z0-9]{3,}$', c_compact):
            return False  # Rejects "SWEDEVAB", "HAMBURG", etc.

    if c_compact.isdigit():
        return 5 <= len(c_compact) <= 20

    return any(ch.isalpha() for ch in c_compact) and any(ch.isdigit() for ch in c_compact)


def extract_bl_number_regex(text: str, current_acid: Optional[str] = None) -> Optional[str]:
    upper = text.upper()
    value_pat = r"([A-Z0-9][A-Z0-9 \-]{3,30}[A-Z0-9])"

    patterns = [
        # Pattern 0: Exact "B/L No." label
        r"B/L\s*No\.?\s*\n?\s*([0-9]+(?: [0-9]+)*)",

        # Pattern 1: M/BL label (master BL from schedule rows above table)
        r"M\s*/?\s*BL\s*[:\-]?\s*([A-Z0-9][A-Z0-9 \-]{2,20})",

        # Pattern 2: ETD line that contains the MBL number
        # e.g. "ETD  46086  COSCO  2311135"
        r"(?:ETD|JOB\s*NO)\.?\s*[:\-]?\s*[A-Z0-9]*\s+([0-9]{4,10})\s+[A-Z]",

        # Standard labels
        rf"(?:BILL\s*OF\s*LADING\s*(?:NO|NUMBER|#)|B/L\s*(?:NO|NUMBER|#)|BL\s*(?:NO|NUMBER|#))\.?\s*[:\-]?\s*\n?\s*{value_pat}",
        rf"\bB/L\s*NO\.?\s*{value_pat}",
        rf"\bBLNO\.?\s*[:\-]?\s*{value_pat}",
        rf"\bREF(?:ERENCE)?\s*(?:NO|NUMBER|#)?\.?\s*[:\-]?\s*{value_pat}",
    ]
    for pat in patterns:
        m = re.search(pat, upper, flags=re.I | re.S)
        if not m:
            continue
        val = m.group(1)
        val = re.split(r"\n| {2,}", val)[0].strip(" ,;:-")
        val_compact = val.replace(" ", "")
        if is_likely_bl_number(val_compact, current_acid):
            if val_compact.isdigit() and " " in val:
                return val
            return val_compact

    header = upper[:2000]
    for val in re.findall(r"\b[A-Z]{2,5}\d[A-Z0-9\-]{4,20}\b|\b\d{5,20}\b", header):
        if is_likely_bl_number(val, current_acid):
            return val
    return None


def extract_acid_regex(text: str) -> Optional[str]:
    upper = text.upper()
    m = re.search(r"\bACID\s*(?:NO|NUMBER)?\.?\s*[:\-]?\s*([0-9][0-9\s\-]{12,30})", upper)
    if m:
        digits = normalize_digits(m.group(1))
        if digits and 10 <= len(digits) <= 19:
            return digits
    # Strict 19-digit fallback
    m = re.search(r"\b(\d{19})\b", upper)
    return m.group(1) if m else None


def extract_hs_code_regex(text: str) -> Optional[str]:
    m = re.search(r"H\.?\s*S\.?\s*[- ]?\s*CODE\s*[:\-]?\s*([0-9\s\-/|]{6,80})", text, flags=re.I)
    if not m:
        return None
    codes = re.findall(r"\b\d{6,10}\b", m.group(1))
    out = []
    for c in codes:
        if c not in out:
            out.append(c)
    return "|".join(out) if out else None


def extract_containers_regex(text: str) -> List[Dict[str, Optional[str]]]:
    containers: List[Dict[str, Optional[str]]] = []
    compact = text.upper()
    for m in re.finditer(r"\b([A-Z]{4}\d{7})\b(?:[/\s\-]+([A-Z0-9]{4,20}))?", compact):
        container = m.group(1)
        seal = m.group(2)
        if container not in [c["container_number"] for c in containers]:
            containers.append({
                "container_number": container,
                "seal_number": seal,
                "container_type": None,
                "packages": None,
                "gross_weight_kg": None,
                "measurement_cbm": None,
            })
    return containers


def extract_vessel_voyage_port_regex(text: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    upper = text.upper()

    # Example: CMA CGM TIGA 0NVLOS1MA HAMBURG
    for port in sorted(KNOWN_PORTS, key=len, reverse=True):
        pat = rf"\b([A-Z][A-Z ]{{3,40}}?)\s+([A-Z0-9]{{5,12}})\s+{re.escape(port)}\b"
        m = re.search(pat, upper)
        if m:
            vessel = clean_value(m.group(1))
            voyage = clean_value(m.group(2))
            return vessel, voyage, port

    # Example: VESSEL: CMA CGM TIGA / 0NVLOS1MA
    m = re.search(r"(?:OCEAN\s+)?VESSEL\s*[:\-]?\s*([A-Z][A-Z0-9 ]{3,50})(?:\s*/\s*|\s+)([A-Z0-9]{4,12})", upper)
    if m:
        return clean_value(m.group(1)), clean_value(m.group(2)), None

    return None, None, None


def infer_direction(data: Dict[str, Any]) -> None:
    if data.get("mesco_direction") is not None:
        return
    dest = (data.get("mesco_destination") or "").upper()
    origin = (data.get("mesco_origin") or "").upper()
    if any(p in dest for p in EGYPT_PORTS):
        data["mesco_direction"] = 300000000
    elif any(p in origin for p in EGYPT_PORTS):
        data["mesco_direction"] = 300000001


def validate_and_correct(data: Dict[str, Any], raw_text: str) -> Dict[str, Any]:
    base = empty_bl_entity()
    base.update(data or {})
    data = base

    for k, v in list(data.items()):
        if isinstance(v, str):
            data[k] = clean_value(v)

    # Regex fallbacks
    if not data.get("mesco_acidnumber"):
        acid = extract_acid_regex(raw_text)
        if acid:
            data["mesco_acidnumber"] = acid
            add_warning(data, "mesco_acidnumber filled by regex fallback.")

    if not data.get("mesco_masterblno"):
        bl = extract_bl_number_regex(raw_text, data.get("mesco_acidnumber"))
        if bl:
            data["mesco_masterblno"] = bl
            add_warning(data, "mesco_masterblno filled by regex fallback.")

    if not data.get("mesco_hscode"):
        hs = extract_hs_code_regex(raw_text)
        if hs:
            data["mesco_hscode"] = hs
            add_warning(data, "mesco_hscode filled by regex fallback.")

    vessel, voyage, port = extract_vessel_voyage_port_regex(raw_text)
    if vessel and not data.get("mesco_vessel"):
        data["mesco_vessel"] = vessel
        add_warning(data, "mesco_vessel filled by regex fallback.")
    if voyage and not data.get("mesco_voytruckno"):
        data["mesco_voytruckno"] = voyage
        add_warning(data, "mesco_voytruckno filled by regex fallback.")
    if port and not data.get("mesco_origin"):
        data["mesco_origin"] = port
        add_warning(data, "mesco_origin inferred from vessel/voyage/port line.")

    found_containers = extract_containers_regex(raw_text)
    if not data.get("containers") and found_containers:
        data["containers"] = found_containers
        add_warning(data, "containers filled by regex fallback.")

    # Normalize ACID and avoid B/L = ACID
    if data.get("mesco_acidnumber"):
        acid_digits = normalize_digits(data["mesco_acidnumber"])
        data["mesco_acidnumber"] = acid_digits if acid_digits and len(acid_digits) >= 10 else None

    if data.get("mesco_masterblno") and data.get("mesco_acidnumber"):
        bl_compact = re.sub(r"\D", "", str(data["mesco_masterblno"]))
        acid = str(data["mesco_acidnumber"])
        if bl_compact and bl_compact == acid:
            alt = extract_bl_number_regex(raw_text, acid)
            if alt and re.sub(r"\D", "", alt) != acid:
                data["mesco_masterblno"] = alt
            else:
                add_warning(data, "B/L number equals ACID number; verify manually.")

    # Normalize numeric totals
    data["cr401_totalgrossweight"] = normalize_numeric(data.get("cr401_totalgrossweight"))
    data["cr401_totalvolume"] = normalize_numeric(data.get("cr401_totalvolume"))

    # Normalize containers
    cleaned_containers: List[Dict[str, Optional[str]]] = []
    for item in data.get("containers") or []:
        if isinstance(item, BaseModel):
            item = item.model_dump()
        if not isinstance(item, dict):
            continue
        c = {
            "container_number": clean_value(item.get("container_number")),
            "seal_number": clean_value(item.get("seal_number")),
            "container_type": clean_value(item.get("container_type")),
            "packages": clean_value(item.get("packages")),
            "gross_weight_kg": normalize_numeric(item.get("gross_weight_kg")),
            "measurement_cbm": normalize_numeric(item.get("measurement_cbm")),
        }
        if c["container_number"]:
            c["container_number"] = c["container_number"].upper().replace(" ", "")
        if any(c.values()):
            cleaned_containers.append(c)

    data["containers"] = cleaned_containers
    if cleaned_containers:
        first = cleaned_containers[0]
        data["container_number"] = data.get("container_number") or first.get("container_number")
        data["seal_number"] = data.get("seal_number") or first.get("seal_number")
        data["mesco_containertype"] = data.get("mesco_containertype") or first.get("container_type")

    # Inferences
    upper = raw_text.upper()
    data["mesco_transporttype"] = data.get("mesco_transporttype") or 300000000

    if data.get("mesco_loadtype") is None:
        if "LCL" in upper:
            data["mesco_loadtype"] = 300000001
        elif cleaned_containers or re.search(r"\bFCL\b", upper):
            data["mesco_loadtype"] = 300000000

    if not data.get("mesco_pcfreightterm"):
        if "FREIGHT COLLECT" in upper:
            data["mesco_pcfreightterm"] = "COLLECT"
        elif "FREIGHT PREPAID" in upper:
            data["mesco_pcfreightterm"] = "PREPAID"

    if "TELEX RELEASE" in upper or "EXPRESS RELEASE" in upper:
        data["mesco_telexrelease"] = True

    infer_direction(data)

    data.setdefault("confidence", {})
    data["confidence"]["post_validation"] = "completed"
    data["confidence"]["bl_number_rule"] = "accepted" if data.get("mesco_masterblno") else "missing"
    data["confidence"]["container_number_rule"] = "accepted" if data.get("container_number") else "missing"

    return BLEntity(**data).model_dump()


# ============================================================
# Routes
# ============================================================

@app.get("/", response_class=HTMLResponse)
async def root() -> str:
    return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>B/L Extractor</title>
  <style>
    body{font-family:Arial,sans-serif;margin:40px;max-width:900px}
    input,button{font-size:16px;padding:8px;margin:6px 0}
    pre{white-space:pre-wrap;background:#f7f7f7;padding:16px;border-radius:8px}
  </style>
</head>
<body>
  <h1>Professional PDF + OCR + Excel Bill of Lading Extractor</h1>
  <p>Upload a PDF, XLSX, XLS, or CSV. PDFs use native coordinate text plus OCR/hybrid when needed; spreadsheets use row/cell extraction.</p>
  <form id="f">
    <input type="file" id="file" name="file" accept=".pdf,.xlsx,.xls,.csv,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,text/csv" required />
    <label><input type="checkbox" id="force_ocr" /> Force OCR/hybrid</label><br/>
    <button type="submit">Extract</button>
  </form>
  <pre id="out"></pre>
<script>
document.getElementById('f').addEventListener('submit', async (e)=>{
  e.preventDefault();
  const fd = new FormData();
  fd.append('file', document.getElementById('file').files[0]);
  const force = document.getElementById('force_ocr').checked ? '?force_ocr=true' : '';
  const res = await fetch('/extract' + force, {method:'POST', body:fd});
  document.getElementById('out').textContent = JSON.stringify(await res.json(), null, 2);
});
</script>
</body>
</html>
"""


@app.get("/health")
async def health() -> Dict[str, Any]:
    return {"status": "ok", "version": app.version}


@app.post("/extract-text")
async def extract_text(file: UploadFile = File(...), force_ocr: bool = False) -> JSONResponse:
    filename = _safe_filename(file.filename)
    file_bytes = await file.read()
    result = extract_document_text_professionally(file_bytes, filename, force_ocr=force_ocr)
    return JSONResponse(result)


@app.post("/extract")
async def extract(file: UploadFile = File(...), force_ocr: bool = False) -> JSONResponse:
    filename = _safe_filename(file.filename)
    file_bytes = await file.read()
    text_result = extract_document_text_professionally(file_bytes, filename, force_ocr=force_ocr)
    if should_return_spreadsheet_record_array(text_result):
        return JSONResponse(extract_spreadsheet_records_with_ai(text_result, filename, include_raw_text=False))

    data = extract_with_azure_openai_mapped(text_result["text"])

    data["source_filename"] = filename
    data["source_file_type"] = text_result.get("source_file_type")
    data["extraction_method"] = text_result["method"]
    data["extraction_quality"] = text_result["quality"]
    return JSONResponse(data)


@app.post("/extract-without-ai")
async def extract_without_ai(file: UploadFile = File(...), force_ocr: bool = False) -> JSONResponse:
    """
    Useful for testing the PDF/native/OCR/spreadsheet text layer without Azure OpenAI.
    It returns deterministic regex fallbacks only, not full semantic extraction.
    """
    filename = _safe_filename(file.filename)
    file_bytes = await file.read()
    text_result = extract_document_text_professionally(file_bytes, filename, force_ocr=force_ocr)
    if should_return_spreadsheet_record_array(text_result):
        debug_records = []
        for record_info in text_result.get("records") or []:
            record_text = record_info.get("text") or ""
            row_data = apply_spreadsheet_record_seed(empty_bl_entity(), record_info, record_text)
            row_data["record_index"] = record_info.get("record_index")
            row_data["source_sheet"] = record_info.get("sheet_name")
            row_data["source_row"] = record_info.get("source_row")
            row_data["crm_payload"] = to_crm_payload(row_data, include_nulls=False)
            debug_records.append(row_data)
        return JSONResponse({
            "source_filename": filename,
            "source_file_type": text_result.get("source_file_type"),
            "extraction_method": text_result.get("method"),
            "extraction_quality": text_result.get("quality"),
            "record_count": len(debug_records),
            "records": debug_records,
        })

    data = validate_and_correct(empty_bl_entity(), text_result["text"])
    data["source_filename"] = filename
    data["source_file_type"] = text_result.get("source_file_type")
    data["extraction_method"] = text_result["method"]
    data["extraction_quality"] = text_result["quality"]
    data["raw_text_preview"] = text_result["text"][:5000]
    return JSONResponse(data)

# ============================================================
# AI-first CRM mapping helpers and endpoint
# ============================================================

CRM_FIELD_ORDER = [
    "document_type",
    "mesco_masterblno",
    "mesco_bookingnumber",
    "mesco_acidnumber",
    "mesco_shippernamecontactno",
    "mesco_shipperaddress",
    "mesco_consigneenamecontactno",
    "mesco_consigneeaddress",
    "mesco_notify1",
    "mesco_notifyaddress",
    "mesco_vessel",
    "mesco_voytruckno",
    "mesco_origin",
    "mesco_destination",
    "mesco_cargodescription",
    "cr401_totalgrossweight",
    "cr401_totalvolume",
    "cr401_totalpackages",
    "mesco_nooforgbls",
    "mesco_containertype",
    "mesco_containertype2",
    "mesco_containertype3",
    "mesco_handlinginformation",
    "mesco_freightpayableat",
    "mesco_ponumber",
    "mesco_customerreference",
    "mesco_bltype",
    "mesco_transporttype",
    "mesco_loadtype",
    "mesco_direction",
    "cr401_totalteus",
    "mesco_pcfreightterm",
    "mesco_etdorigin",
    "mesco_etadestination",
    "mesco_pickupaddress",
    "mesco_deliveryaddress",
    "mesco_transhipmentport",
    "mesco_importerstaxno",
    "mesco_foreignsupplierregistrationnumber",
    "mesco_incoterm",
    "mesco_telexrelease",
    "mesco_imoclass",
    "mesco_unnumber",
    "mesco_hscode",
    "mesco_dateofissue",
    "mesco_placeofissue",
    "mesco_shippedonboarddate",
]

FIELD_DESCRIPTIONS_FOR_AI = {
    "mesco_masterblno": "Bill of Lading number only. Label examples: BILL OF LADING NO, B/L No., BL NO. Never use booking, ACID, container, postal code, company name.",
    "mesco_bookingnumber": "Booking number only. Label examples: BOOKING NO, Booking Ref.",
    "mesco_acidnumber": "Egypt ACID number. Return digits only, normally 19 digits but accept visible shorter/longer only when explicitly labeled ACID.",
    "mesco_shippernamecontactno": "Shipper company/name and contact line only, not full address unless no separation is possible.",
    "mesco_shipperaddress": "Full shipper address.",
    "mesco_consigneenamecontactno": "Consignee company/name and contact line only, not full address unless no separation is possible.",
    "mesco_consigneeaddress": "Full consignee address.",
    "mesco_notify1": "Notify party name/company.",
    "mesco_notifyaddress": "Notify party address.",
    "mesco_vessel": "Vessel name only. Do not include voyage or port.",
    "mesco_voytruckno": "Voyage number/truck number only.",
    "mesco_origin": "Port of loading/origin.",
    "mesco_destination": "Port of discharge/destination.",
    "mesco_cargodescription": "Actual cargo/goods description. Exclude legal boilerplate unless it is part of cargo description.",
    "cr401_totalgrossweight": "Total gross weight in kg as numeric string where possible.",
    "cr401_totalvolume": "Total measurement/volume in CBM as numeric string where possible.",
    "cr401_totalpackages": "Total packages / number of packages.",
    "mesco_nooforgbls": "Number of original B/Ls.",
    "mesco_containertype": "Primary container type, e.g. 40 HC, 20 GP, 40 High Cube.",
    "mesco_pcfreightterm": "PREPAID or COLLECT.",
    "mesco_freightpayableat": "Freight payable at location/city.",
    "mesco_transporttype": "Use 300000000 for sea shipments.",
    "mesco_loadtype": "Use 300000000 for FCL/full container; 300000001 for LCL.",
    "mesco_direction": "Use 300000000 for import to Egypt; 300000001 for export from Egypt.",
    "mesco_bltype": "Use 886150001 only for Original B/L if ORIGINAL is actually stamped/issued. Non-negotiable is not original.",
    "mesco_telexrelease": "True only if TELEX RELEASE or EXPRESS RELEASE is visible.",
    "mesco_hscode": "HS code(s), 6-10 digits, pipe-separated if multiple.",
}


def build_ai_evidence_pack(text: str) -> Dict[str, Any]:
    """Provide the model with deterministic candidates so it maps fields correctly."""
    bl = extract_bl_number_regex(text)
    acid = extract_acid_regex(text)
    vessel, voyage, bled_origin = extract_vessel_voyage_port_regex(text)
    containers = extract_containers_regex(text)
    hs = extract_hs_code_regex(text)
    upper = text.upper()

    ports_seen = [p for p in sorted(KNOWN_PORTS, key=len, reverse=True) if p in upper]

    return {
        "candidate_bl_number": bl,
        "candidate_acid_number": acid,
        "candidate_hs_code": hs,
        "candidate_vessel": vessel,
        "candidate_voyage": voyage,
        "candidate_origin_from_vessel_line": bled_origin,
        "candidate_containers": containers,
        "ports_seen_in_text": ports_seen[:20],
        "freight_collect_seen": "FREIGHT COLLECT" in upper,
        "freight_prepaid_seen": "FREIGHT PREPAID" in upper,
        "telex_release_seen": "TELEX RELEASE" in upper or "EXPRESS RELEASE" in upper,
        "non_negotiable_seen": "NON-NEGOTIABLE" in upper or "NON NEGOTIABLE" in upper,
        "original_seen": "ORIGINAL" in upper,
    }


def to_crm_payload(data: Dict[str, Any], include_nulls: bool = False) -> Dict[str, Any]:
    """
    Final Dynamics/Mesco CRM payload.
    Keeps only CRM fields in the correct names/order and removes internal extraction metadata.
    """
    payload: Dict[str, Any] = {}
    for field in CRM_FIELD_ORDER:
        value = data.get(field)
        if include_nulls or value not in (None, "", [], {}):
            payload[field] = value

    # Keep top-level primary container fields for existing CRM integrations.
    for field in ("container_number", "seal_number", "containers"):
        value = data.get(field)
        if include_nulls or value not in (None, "", [], {}):
            payload[field] = value

    return payload


AI_MAPPING_SYSTEM_PROMPT = SYSTEM_PROMPT + """

CRM MAPPING MODE:
- You are mapping directly to Mesco CRM field names.
- Every output key must be one of the JSON schema keys. Do not invent user-facing labels.
- Prefer explicit field labels over guesses.
- Use the deterministic evidence pack only as hints; the extracted document text remains the source of truth.
- If a candidate conflicts with a clearly labeled value in the extracted text, use the clearly labeled value.
- For B/L number, prefer the value after BILL OF LADING NO / B/L NO / HBL NO / MBL NO over every other candidate.
- For Excel sheets, use column headers and the same row's values first, then nearby key/value pairs.
- For the sample SACO style: BOOKING NO is booking, BILL OF LADING NO is mesco_masterblno, EXPORT REFERENCES can be customer reference.
- For rows like "CMA CGM TIGA 0NVLOS1MA HAMBURG": vessel=CMA CGM TIGA, voyage=0NVLOS1MA, origin=HAMBURG.
- For rows like "PORT SAID PORT SAID WEST": destination/port of discharge=PORT SAID and delivery=PORT SAID WEST when labels support that.
- Container syntax like "TCLU5466216/B135553" means container_number=TCLU5466216 and seal_number=B135553.

Field meanings:
""" + json.dumps(FIELD_DESCRIPTIONS_FOR_AI, indent=2)


def extract_with_azure_openai_mapped(extracted_text: str) -> Dict[str, Any]:
    """
    AI-first extraction with deterministic evidence and final CRM mapping validation.
    This is the recommended function used by /extract-ai-full.
    """
    text = normalize_text(extracted_text)
    if not text:
        raise HTTPException(status_code=400, detail="No text was extracted from the PDF.")

    evidence = build_ai_evidence_pack(text)
    client = get_azure_client()
    safe_text = text[: settings.max_input_chars]

    user_prompt = (
        "Extract and map this Bill of Lading into the exact Mesco CRM JSON schema.\n\n"
        "DETERMINISTIC EVIDENCE PACK (hints, not a replacement for the text):\n"
        + json.dumps(evidence, ensure_ascii=False, indent=2)
        + "\n\nPDF EXTRACTED TEXT:\n"
        + safe_text
    )

    try:
        response = client.chat.completions.create(
            model=settings.azure_openai_deployment,
            temperature=0,
            messages=[
                {"role": "system", "content": AI_MAPPING_SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_schema", "json_schema": JSON_SCHEMA},
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI mapped extraction failed: {exc}")

    content = response.choices[0].message.content
    if not content:
        raise HTTPException(status_code=502, detail="Azure OpenAI returned empty content.")

    try:
        raw_ai = json.loads(content)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI returned invalid JSON: {exc}")

    validated = validate_and_correct(raw_ai, extracted_text)
    validated["ai_evidence"] = evidence
    validated["crm_payload"] = to_crm_payload(validated, include_nulls=False)
    validated["crm_payload_with_nulls"] = to_crm_payload(validated, include_nulls=True)
    validated.setdefault("confidence", {})["ai_mapping_mode"] = "enabled"
    return validated


def _record_header_value(record_info: Dict[str, Any], *header_patterns: str) -> Optional[str]:
    values = record_info.get("values_by_header") or {}
    for header, value in values.items():
        if value in (None, ""):
            continue
        for pattern in header_patterns:
            if re.search(pattern, header, flags=re.I):
                return clean_value(value)
    return None


def _first_incoterm(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    m = re.search(r"\b(EXW|FOB|FCA|CIF|CFR|CPT|CIP|DAP|DDP|DAT|DPU)\b", value.upper())
    return m.group(1) if m else None


def build_spreadsheet_record_seed(record_info: Dict[str, Any]) -> Dict[str, Any]:
    """Deterministic row-level mapping to prevent AI from mixing Excel rows."""
    cargo = _record_header_value(record_info, r"CARGO\s*TYPE", r"CARGO", r"DESCRIPTION")
    remarks = _record_header_value(record_info, r"REMARK") or ""
    delivery_term = _record_header_value(record_info, r"DELIVERY\s*TERM", r"INCOTERM")

    # H/BL ACID from its own dedicated column — exact match first
    acid = (
        _record_header_value(record_info, r"^H/BL\s*ACID$", r"H\s*/?\s*BL\s*ACID")
        or extract_acid_regex(cargo or "")
    )

    # H/BL number from the H/BL Nos. column — must NOT match H/BL ACID column
    hbl_no = _record_header_value(record_info, r"^H/BL\s*Nos?\.$", r"^H/BL\s*Nos?", r"^HBL")

    # Shipper: the full cell value may contain name + address merged — split on first newline or ADDRESS:
    shipper_raw = _record_header_value(record_info, r"^SHIPPER$", r"SHIPPER") or ""
    shipper_name = re.split(r"\s+ADDRESS\s*:", shipper_raw, maxsplit=1)[0].strip() if shipper_raw else None
    shipper_address = re.split(r"\s+ADDRESS\s*:", shipper_raw, maxsplit=1)[1].strip() if shipper_raw and "ADDRESS" in shipper_raw.upper() else None

    # Consignee: same split approach
    consignee_raw = _record_header_value(record_info, r"^CONSIGNEE$", r"CONSIGNEE") or ""
    # Extract tax number from consignee field
    tax_match = re.search(r"TAX\s*(?:NO|ID)\.?\s*[:\-]?\s*([0-9]{6,15})", consignee_raw, re.I)
    importer_tax = tax_match.group(1) if tax_match else None
    # Consignee name is typically the first word(s) before the address
    consignee_name = re.split(r"\s{2,}|\n", consignee_raw)[0].strip() if consignee_raw else None
    consignee_address = consignee_raw  # keep full for address field

    # Extract M/BL-level fields from global context rows (above the table)
    record_text = record_info.get("text") or ""

    # M/BL number — from global context, labeled "M/BL" or after ETD column
    mbl_match = re.search(r"M\s*/?\s*BL\s*[:\-]?\s*([A-Z0-9]{4,15})\b", record_text, re.I)
    master_bl = mbl_match.group(1) if mbl_match else None

    # ETD — the actual ETD value (46086 is voyage/schedule number, NOT ETD date)
    # ETD column in this file is blank; leave null rather than capturing wrong value
    etd_match = re.search(
        r"ETD\s*[:\-]?\s*(\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        record_text, re.I
    )
    etd_value = etd_match.group(1) if etd_match else None

    # Carrier from global context
    carrier_match = re.search(
        r"\b(COSCO|MSC|MAERSK|CMA\s*CGM|HAPAG|EVERGREEN|YANG\s*MING|ONE|HMM|ZIM|PIL|ARKAS)\b",
        record_text, re.I,
    )
    carrier = carrier_match.group(1).upper() if carrier_match else None

    # Job number from global context
    job_match = re.search(r"JOB\s*NO\.?\s*[:\-]?\s*([A-Z0-9]{5,12})", record_text, re.I)
    job_no = job_match.group(1) if job_match else None

    # M/BL ACID from global context
    master_acid_match = re.search(
        r"M\s*/?\s*BL\s*ACID\s*[:\-]?\s*([0-9]{10,19})", record_text, re.I
    )
    master_acid = master_acid_match.group(1) if master_acid_match else None

    seed: Dict[str, Any] = {
        "document_type": "Bill of Lading",
        # H/BL number goes to mesco_masterblno (each row IS its own BL)
        "mesco_masterblno": hbl_no,
        # Job number goes to customer reference, NOT masterblno
        "mesco_customerreference": job_no,
        "mesco_acidnumber": acid or master_acid,
        "mesco_shippernamecontactno": shipper_name or shipper_raw or None,
        "mesco_shipperaddress": shipper_address,
        "mesco_consigneenamecontactno": consignee_name,
        "mesco_consigneeaddress": consignee_address,
        "mesco_importerstaxno": importer_tax,
        "mesco_origin": _record_header_value(record_info, r"PLACE\s*OF\s*RECEIPT", r"POL", r"ORIGIN"),
        "mesco_destination": _record_header_value(record_info, r"PLACE\s*OF\s*DELIVERY", r"POD", r"DESTINATION"),
        # Per-row package/weight/volume — NOT the sheet total row
        "cr401_totalpackages": _record_header_value(record_info, r"NOS\.?\s*OF\s*PACK", r"PACKAGE"),
        "cr401_totalgrossweight": normalize_numeric(
            _record_header_value(record_info, r"GROSS\s*WEIGHT", r"GROSS", r"WEIGHT")
        ),
        "cr401_totalvolume": normalize_numeric(
            _record_header_value(record_info, r"MEASUR", r"CBM", r"VOLUME")
        ),
        "mesco_cargodescription": cargo,
        "mesco_hscode": extract_hs_code_regex(cargo or ""),
        "mesco_incoterm": _first_incoterm(delivery_term),
        # ETD — only set if a real date was found, not the schedule number
        "mesco_etdorigin": etd_value,
        "mesco_transporttype": 300000000,
        "mesco_loadtype": 300000001,  # LCL loading sheet
    }

    if remarks:
        if "TELEX" in remarks.upper() or "EXPRESS" in remarks.upper():
            seed["mesco_telexrelease"] = True
        if "ORIGINAL" in remarks.upper():
            seed["mesco_bltype"] = 886150001

    return {k: v for k, v in seed.items() if v not in (None, "", [], {})}


def apply_spreadsheet_record_seed(data: Dict[str, Any], record_info: Dict[str, Any], record_text: str) -> Dict[str, Any]:
    seed = build_spreadsheet_record_seed(record_info)
    override_fields = {
        "document_type", "mesco_masterblno", "mesco_acidnumber", "mesco_shippernamecontactno",
        "mesco_shipperaddress", "mesco_consigneenamecontactno", "mesco_consigneeaddress",
        "mesco_importerstaxno", "mesco_pickupaddress", "mesco_deliveryaddress",
        "mesco_origin", "mesco_destination", "cr401_totalpackages", "cr401_totalgrossweight",
        "cr401_totalvolume", "mesco_cargodescription", "mesco_hscode", "mesco_incoterm",
        "mesco_telexrelease", "mesco_bltype", "mesco_transporttype", "mesco_loadtype",
        "mesco_customerreference", "mesco_etdorigin",
    }
    for field, value in seed.items():
        if field in override_fields or not data.get(field):
            data[field] = value

    data = validate_and_correct(data, record_text)

    # Suppress vessel/voyage — spreadsheet address text triggers false positives
    # Only keep if there is an actual VESSEL column header in this row
    values = record_info.get("values_by_header") or {}
    has_vessel_header = any(re.search(r"VESSEL|VOYAGE", h, re.I) for h in values)
    if not has_vessel_header:
        data["mesco_vessel"] = None
        data["mesco_voytruckno"] = None

    # Re-apply seed after validation to prevent validate_and_correct() overwriting correct row values
    for field, value in seed.items():
        if field in override_fields:
            data[field] = value  # always force row-level values, no "or not data.get"

    data.setdefault("confidence", {})["spreadsheet_row_seed_applied"] = True
    return data


def extract_spreadsheet_records_with_ai(text_result: Dict[str, Any], filename: str, include_raw_text: bool = False) -> Dict[str, Any]:
    records_meta = text_result.get("records") or []
    extracted_records: List[Dict[str, Any]] = []

    for record_info in records_meta:
        record_text = record_info.get("text") or ""
        if not record_text.strip():
            continue
        data = extract_with_azure_openai_mapped(record_text)
        data = apply_spreadsheet_record_seed(data, record_info, record_text)
        data["record_index"] = record_info.get("record_index")
        data["source_sheet"] = record_info.get("sheet_name")
        data["source_row"] = record_info.get("source_row")
        data["source_header_row"] = record_info.get("header_row")
        data["source_filename"] = filename
        data["source_file_type"] = "spreadsheet"
        data["extraction_method"] = text_result.get("method")
        data["crm_payload"] = to_crm_payload(data, include_nulls=False)
        data["crm_payload_with_nulls"] = to_crm_payload(data, include_nulls=True)
        if include_raw_text:
            data["raw_record_text"] = record_text
        extracted_records.append(data)

    return {
        "document_type": "Bill of Lading",
        "source_filename": filename,
        "source_file_type": "spreadsheet",
        "extraction_method": text_result.get("method"),
        "extraction_quality": text_result.get("quality", {}),
        "record_count": len(extracted_records),
        "records": extracted_records,
        "crm_payloads": [r.get("crm_payload", {}) for r in extracted_records],
        "crm_payloads_with_nulls": [r.get("crm_payload_with_nulls", {}) for r in extracted_records],
        "warnings": [] if extracted_records else ["No spreadsheet table records were detected."],
    }


def should_return_spreadsheet_record_array(text_result: Dict[str, Any]) -> bool:
    return text_result.get("source_file_type") == "spreadsheet" and len(text_result.get("records") or []) > 1



@app.post("/extract-ai-full")
async def extract_ai_full(
    file: UploadFile = File(...),
    force_ocr: bool = True,
    include_raw_text: bool = False,
) -> JSONResponse:
    """
    Recommended endpoint.
    - Supports PDF, XLSX, XLS, and CSV.
    - Uses native + OCR hybrid by default for PDFs and row/cell extraction for spreadsheets.
    - Sends the complete extracted text to Azure OpenAI.
    - For spreadsheets with multiple table rows, returns records[] plus crm_payloads[].
    - Returns validated Mesco CRM field mapping plus crm_payload.
    """
    filename = _safe_filename(file.filename)
    file_bytes = await file.read()
    text_result = extract_document_text_professionally(file_bytes, filename, force_ocr=force_ocr)
    if should_return_spreadsheet_record_array(text_result):
        return JSONResponse(extract_spreadsheet_records_with_ai(text_result, filename, include_raw_text=include_raw_text))

    data = extract_with_azure_openai_mapped(text_result["text"])
    data["source_filename"] = filename
    data["source_file_type"] = text_result.get("source_file_type")
    data["extraction_method"] = text_result["method"]
    data["extraction_quality"] = text_result["quality"]
    if include_raw_text:
        data["raw_text"] = text_result["text"]
    return JSONResponse(data)
