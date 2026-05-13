import csv
import io
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from config import settings
from pdf_extractor import count_field_hits, normalize_text


SUPPORTED_SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}


def _file_extension(filename: Optional[str]) -> str:
    return Path(filename or "").suffix.lower()


def normalize_spreadsheet_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.time() else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return (f"{value:.10f}").rstrip("0").rstrip(".")
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > settings.excel_max_cell_chars:
        text = text[: settings.excel_max_cell_chars].rstrip() + "…"
    return text


def _is_meaningful_spreadsheet_row(values: List[str]) -> bool:
    return any(v.strip() for v in values)


def _row_has_business_terms(values: List[str]) -> bool:
    row_text = " ".join(values).upper()
    terms = (
        "B/L", "BL", "BILL OF LADING", "BOOKING", "MBL", "HBL", "HOUSE", "MASTER",
        "CONTAINER", "SEAL", "SHIPPER", "CONSIGNEE", "NOTIFY", "VESSEL", "VOYAGE",
        "POL", "POD", "PORT", "ORIGIN", "DESTINATION", "GROSS", "WEIGHT", "CBM",
        "MEAS", "PACKAGE", "PCS", "CTNS", "HS", "ACID", "FREIGHT", "PREPAID", "COLLECT",
        "INVOICE", "MANIFEST", "ETA", "ETD", "CFS", "LCL", "FCL",
    )
    return any(term in row_text for term in terms)


def _column_name(index_1_based: int) -> str:
    name = ""
    n = index_1_based
    while n:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def _xlsx_col_to_index(col: str) -> int:
    n = 0
    for ch in col.upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
    return n


def _xlsx_cell_ref_to_row_col(ref: str) -> Tuple[int, int]:
    m = re.match(r"([A-Z]+)(\d+)", ref.upper())
    if not m:
        return 1, 1
    return int(m.group(2)), _xlsx_col_to_index(m.group(1))


def _xml_text(element: Optional[ET.Element]) -> str:
    if element is None:
        return ""
    return "".join(element.itertext())


def _detect_key_value_pairs_from_grid(grid: List[List[str]], sheet_name: str, max_pairs: int = 180) -> List[str]:
    pairs: List[str] = []
    label_regex = re.compile(
        r"\b(B/L|BL|BILL OF LADING|BOOKING|MBL|HBL|SHIPPER|CONSIGNEE|NOTIFY|VESSEL|VOYAGE|"
        r"POL|POD|PORT|ORIGIN|DESTINATION|CONTAINER|SEAL|GROSS|WEIGHT|CBM|MEAS|PACKAGE|"
        r"ACID|HS|FREIGHT|ETA|ETD|INVOICE|REF|REFERENCE|TERMS)\b",
        re.I,
    )
    for r_idx, row in enumerate(grid):
        if len(pairs) >= max_pairs:
            break
        for c_idx, cell in enumerate(row):
            if not cell or not label_regex.search(cell):
                continue
            candidates: List[str] = []
            for k in range(c_idx + 1, min(len(row), c_idx + 5)):
                if row[k]:
                    candidates.append(row[k])
            for rr in range(r_idx + 1, min(len(grid), r_idx + 5)):
                if c_idx < len(grid[rr]) and grid[rr][c_idx]:
                    candidates.append(grid[rr][c_idx])
            if candidates:
                address = f"{_column_name(c_idx + 1)}{r_idx + 1}"
                pairs.append(f"{sheet_name}!{address}: {cell} => {' | '.join(candidates[:4])}")
    return pairs


def _grid_to_ai_text(sheet_name: str, grid: List[List[str]], max_cols: int) -> str:
    lines: List[str] = []
    non_empty_rows = 0
    business_rows = 0
    for row_no, row in enumerate(grid, start=1):
        trimmed = row[:max_cols]
        if not _is_meaningful_spreadsheet_row(trimmed):
            continue
        non_empty_rows += 1
        if _row_has_business_terms(trimmed):
            business_rows += 1
        row_values = "\t".join(trimmed).rstrip()
        addressed = " | ".join(
            f"{_column_name(i + 1)}{row_no}={v}"
            for i, v in enumerate(trimmed)
            if v
        )
        lines.append(f"ROW {row_no}: {row_values}")
        if addressed:
            lines.append(f"CELLS {row_no}: {addressed}")
    if not lines:
        lines.append("[empty sheet]")
    return "\n".join([
        f"--- SHEET: {sheet_name} ---",
        f"Non-empty rows included: {non_empty_rows}; business-keyword rows: {business_rows}",
        *lines,
    ])


SPREADSHEET_RECORD_HEADER_TERMS = (
    "H/BL", "HBL", "HOUSE", "B/L", "SHIPPER", "CONSIGNEE", "PLACE OF RECEIPT",
    "PLACE OF DELIVERY", "POL", "POD", "PACKAGE", "PACKING", "GROSS", "WEIGHT",
    "MEAS", "CBM", "CARGO", "DESCRIPTION", "ACID", "HS", "REMARK",
    "STATUS", "TERM", "DELIVERY TERM", "NOMINATED", "FREE HAND", "GATE",
    "NOS. OF", "NOS OF", "MEASUREM", "CARGO TYPE", "H/BL ACID", "RATE",
)


def _normalize_header_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (value or "").upper())


def _spreadsheet_header_score(row: List[str]) -> int:
    score = 0
    seen = set()
    for cell in row:
        up = (cell or "").upper()
        if not up:
            continue
        for term in SPREADSHEET_RECORD_HEADER_TERMS:
            if term in up and term not in seen:
                score += 1
                seen.add(term)
    return score


def _looks_like_total_or_summary_row(row: List[str]) -> bool:
    row_text = " ".join(v for v in row if v).strip().upper()
    if not row_text:
        return True
    if re.search(r"^\s*TOTAL\b", row_text):
        return True
    first = next((v.strip().upper() for v in row if v.strip()), "")
    if re.match(r"^TOTAL\s*:?$|^SUBTOTAL\s*:?$|^GRAND\s+TOTAL\s*:?$", first):
        return True
    if first in {"TOTAL", "SUBTOTAL", "GRAND TOTAL"}:
        return True
    if re.fullmatch(r"TOTAL\s*:?(?:\s+[0-9.,]+)*", row_text):
        return True
    return False


def _global_context_from_grid(sheet_name: str, grid: List[List[str]], header_row_index_0: int, max_lines: int = 12) -> List[str]:
    lines: List[str] = []
    for row_no, row in enumerate(grid[:header_row_index_0], start=1):
        if not _is_meaningful_spreadsheet_row(row):
            continue
        row_values = "\t".join(row).rstrip()
        addressed = " | ".join(f"{_column_name(i + 1)}{row_no}={v}" for i, v in enumerate(row) if v)
        if row_values:
            lines.append(f"{sheet_name} ROW {row_no}: {row_values}")
        if addressed:
            lines.append(f"{sheet_name} CELLS {row_no}: {addressed}")
        if len(lines) >= max_lines:
            break
    return lines


def _is_data_row_not_header(row: List[str]) -> bool:
    long_cells = sum(1 for v in row if len(v) > 40)
    return long_cells >= 2


def _extract_table_records_from_grid(sheet_name: str, grid: List[List[str]], max_records: int = 300) -> List[Dict[str, Any]]:
    if not grid:
        return []

    header_candidates = []
    for idx, row in enumerate(grid):
        score = _spreadsheet_header_score(row)
        if score >= 3:
            headers = [h.strip() for h in row]
            non_empty_header_count = sum(1 for h in headers if h and not re.match(r'^列\d+$', h))
            if non_empty_header_count >= 4 and not _is_data_row_not_header(row):
                header_candidates.append((idx, score))

    if not header_candidates:
        return []

    records: List[Dict[str, Any]] = []

    for section_num, (best_idx, _score) in enumerate(header_candidates):
        headers = [h.strip() for h in grid[best_idx]]
        global_context = _global_context_from_grid(sheet_name, grid, best_idx)

        data_start = best_idx + 1
        data_end = len(grid)
        if section_num + 1 < len(header_candidates):
            data_end = header_candidates[section_num + 1][0]

        for row_idx in range(data_start, data_end):
            row = grid[row_idx]
            if not _is_meaningful_spreadsheet_row(row):
                continue
            if _looks_like_total_or_summary_row(row):
                continue

            values_by_header: Dict[str, str] = {}
            filled_cells = 0
            for col_idx, header in enumerate(headers):
                if not header or col_idx >= len(row):
                    continue
                value = row[col_idx].strip()
                if not value:
                    continue
                values_by_header[header] = value
                filled_cells += 1

            row_text_upper = " ".join(row).upper()
            has_record_identifier = bool(
                re.search(r"\b[A-Z]{2,}[A-Z0-9]*\d{4,}[A-Z0-9]*\b", row_text_upper)
                or re.search(r"\b\d{10,19}\b", row_text_upper)
                or any(h and v for h, v in values_by_header.items() if re.search(r"H\s*/?\s*BL|HOUSE|B/L|BL\s*NO", h, re.I))
            )
            has_party = any(re.search(r"SHIPPER|CONSIGNEE", h, re.I) for h in values_by_header)

            if filled_cells < 3 or not (has_record_identifier or has_party):
                continue

            record_index = len(records) + 1
            header_line = "\t".join(headers).rstrip()
            row_line = "\t".join(row[:len(headers)]).rstrip()
            mapped_lines = []
            for header, value in values_by_header.items():
                col_idx = headers.index(header)
                mapped_lines.append(f"{_column_name(col_idx + 1)}{row_idx + 1} | {header} => {value}")

            record_text = normalize_text("\n".join([
                "[SPREADSHEET SINGLE RECORD]",
                f"Sheet: {sheet_name}",
                f"Record index: {record_index}",
                f"Source Excel row: {row_idx + 1}",
                "Instruction: Extract ONLY this one spreadsheet row as ONE Mesco CRM shipment/B/L payload. Do not mix values from other data rows.",
                "[GLOBAL CONTEXT FROM ROWS ABOVE TABLE]",
                *(global_context or ["No global context detected."]),
                "[TABLE HEADER]",
                f"ROW {best_idx + 1}: {header_line}",
                "[THIS RECORD ROW]",
                f"ROW {row_idx + 1}: {row_line}",
                "[HEADER TO VALUE MAP FOR THIS RECORD]",
                *mapped_lines,
            ]))

            records.append({
                "record_index": record_index,
                "sheet_name": sheet_name,
                "header_row": best_idx + 1,
                "source_row": row_idx + 1,
                "values_by_header": values_by_header,
                "text": record_text,
            })
            if len(records) >= max_records:
                break

        if len(records) >= max_records:
            break

    return records


def extract_xlsx_text_zipxml(excel_bytes: bytes, filename: str, openpyxl_error: Optional[Exception] = None) -> Dict[str, Any]:
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    zf = zipfile.ZipFile(io.BytesIO(excel_bytes))

    with zf:
        names = set(zf.namelist())
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in names:
            try:
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for si in root.findall(f"{ns_main}si"):
                    shared_strings.append(_xml_text(si))
            except Exception:
                shared_strings = []

        rel_map: Dict[str, str] = {}
        if "xl/_rels/workbook.xml.rels" in names:
            rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            for rel in rel_root.findall(f"{rel_ns}Relationship"):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target", "")
                if rid and "worksheet" in target:
                    if not target.startswith("xl/"):
                        target = "xl/" + target.lstrip("/")
                    rel_map[rid] = target

        sheets: List[Tuple[str, str]] = []
        if "xl/workbook.xml" in names:
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            sheets_node = wb_root.find(f"{ns_main}sheets")
            if sheets_node is not None:
                for sheet in sheets_node.findall(f"{ns_main}sheet"):
                    sheet_name = sheet.attrib.get("name", "Sheet")
                    rid = sheet.attrib.get(f"{ns_rel}id")
                    path = rel_map.get(rid or "")
                    if path and path in names:
                        sheets.append((sheet_name, path))

        if not sheets:
            sheet_paths = sorted(n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
            sheets = [(f"Sheet{i}", path) for i, path in enumerate(sheet_paths, start=1)]

        workbook_parts: List[str] = [
            "[SPREADSHEET TEXT]",
            f"Filename: {filename}",
            f"Workbook sheets: {', '.join(name for name, _ in sheets)}",
            "Reader: XLSX XML fallback",
        ]
        if openpyxl_error is not None:
            workbook_parts.append(f"Openpyxl warning: {str(openpyxl_error)[:400]}")

        sheet_summaries: List[Dict[str, Any]] = []
        total_pairs = 0
        all_records: List[Dict[str, Any]] = []

        for sheet_name, sheet_path in sheets:
            try:
                root = ET.fromstring(zf.read(sheet_path))
            except Exception as exc:
                workbook_parts.append(f"--- SHEET: {sheet_name} ---\n[Could not parse sheet XML: {exc}]")
                continue

            rows_map: Dict[int, Dict[int, str]] = {}
            formulas: List[str] = []
            sheet_data = root.find(f"{ns_main}sheetData")
            if sheet_data is not None:
                for row in sheet_data.findall(f"{ns_main}row"):
                    for cell in row.findall(f"{ns_main}c"):
                        ref = cell.attrib.get("r", "A1")
                        r_idx, c_idx = _xlsx_cell_ref_to_row_col(ref)
                        if r_idx > settings.excel_max_rows_per_sheet or c_idx > settings.excel_max_cols_per_sheet:
                            continue
                        cell_type = cell.attrib.get("t")
                        v_node = cell.find(f"{ns_main}v")
                        f_node = cell.find(f"{ns_main}f")
                        is_node = cell.find(f"{ns_main}is")
                        value = ""
                        if cell_type == "s" and v_node is not None:
                            try:
                                idx = int(v_node.text or "0")
                                value = shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
                            except Exception:
                                value = v_node.text or ""
                        elif cell_type == "inlineStr":
                            value = _xml_text(is_node)
                        elif cell_type == "b" and v_node is not None:
                            value = "TRUE" if (v_node.text or "") == "1" else "FALSE"
                        elif v_node is not None:
                            value = v_node.text or ""
                        elif f_node is not None:
                            value = "=" + (f_node.text or "")
                        value = normalize_spreadsheet_cell(value)
                        if value:
                            rows_map.setdefault(r_idx, {})[c_idx] = value
                        if f_node is not None and v_node is not None and len(formulas) < 80:
                            formulas.append(f"{ref}: ={f_node.text or ''} => {normalize_spreadsheet_cell(v_node.text)}")

            max_row = min(max(rows_map.keys(), default=1), settings.excel_max_rows_per_sheet)
            max_col = min(max((max(cols.keys()) for cols in rows_map.values()), default=1), settings.excel_max_cols_per_sheet)
            grid: List[List[str]] = []
            for r in range(1, max_row + 1):
                row_map = rows_map.get(r, {})
                grid.append([row_map.get(c, "") for c in range(1, max_col + 1)])

            records = _extract_table_records_from_grid(sheet_name, grid)
            all_records.extend(records)
            pairs = _detect_key_value_pairs_from_grid(grid, sheet_name)
            total_pairs += len(pairs)
            workbook_parts.append(_grid_to_ai_text(sheet_name, grid, max_col))
            if records:
                workbook_parts.append("[DETECTED SPREADSHEET RECORDS]\n" + "\n".join(
                    f"Record {r['record_index']} from {r['sheet_name']} row {r['source_row']}: "
                    + " | ".join(f"{k}={v}" for k, v in list(r['values_by_header'].items())[:8])
                    for r in records[:50]
                ))
            if pairs:
                workbook_parts.append("[DETECTED KEY/VALUE PAIRS]\n" + "\n".join(pairs[:120]))
            if formulas:
                workbook_parts.append("[FORMULA RESULTS]\n" + "\n".join(formulas[:80]))
            sheet_summaries.append({
                "sheet_name": sheet_name,
                "max_row_seen": max_row,
                "max_col_seen": max_col,
                "non_empty_rows": sum(1 for row in grid if _is_meaningful_spreadsheet_row(row)),
                "detected_key_value_pairs": len(pairs),
                "reader": "xlsx_xml_fallback",
            })

    text = normalize_text("\n\n".join(workbook_parts))
    return {
        "method": "excel_xlsx_xml_fallback",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": len(sheets),
            "sheets": sheet_summaries,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_key_value_pairs": total_pairs,
            "detected_record_count": len(all_records),
            "openpyxl_fallback_reason": str(openpyxl_error)[:500] if openpyxl_error is not None else None,
        },
        "records": all_records,
    }


def extract_xlsx_text(excel_bytes: bytes, filename: str) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("Missing dependency for .xlsx files. Install: pip install openpyxl") from exc

    try:
        wb_values = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=False)
        wb_formulas = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False, read_only=False)
    except Exception as exc:
        return extract_xlsx_text_zipxml(excel_bytes, filename, openpyxl_error=exc)

    workbook_parts: List[str] = [
        "[SPREADSHEET TEXT]",
        f"Filename: {filename}",
        f"Workbook sheets: {', '.join(wb_values.sheetnames)}",
        "Instructions for AI: Values are presented with row numbers and cell addresses. Use explicit labels and nearby values.",
    ]
    all_pairs: List[str] = []
    all_records: List[Dict[str, Any]] = []
    sheet_summaries: List[Dict[str, Any]] = []

    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        ws_formula = wb_formulas[sheet_name]
        max_row = min(ws.max_row or 1, settings.excel_max_rows_per_sheet)
        max_col = min(ws.max_column or 1, settings.excel_max_cols_per_sheet)
        grid: List[List[str]] = []
        formula_notes: List[str] = []

        for r in range(1, max_row + 1):
            row_vals: List[str] = []
            for c in range(1, max_col + 1):
                value = normalize_spreadsheet_cell(ws.cell(r, c).value)
                formula_value = ws_formula.cell(r, c).value
                if isinstance(formula_value, str) and formula_value.startswith("=") and value:
                    if len(formula_notes) < 80:
                        formula_notes.append(f"{_column_name(c)}{r}: {formula_value} => {value}")
                row_vals.append(value)
            grid.append(row_vals)

        records = _extract_table_records_from_grid(sheet_name, grid)
        all_records.extend(records)
        pairs = _detect_key_value_pairs_from_grid(grid, sheet_name)
        all_pairs.extend(pairs)
        workbook_parts.append(_grid_to_ai_text(sheet_name, grid, max_col))
        if records:
            workbook_parts.append("[DETECTED SPREADSHEET RECORDS]\n" + "\n".join(
                f"Record {r['record_index']} from {r['sheet_name']} row {r['source_row']}: "
                + " | ".join(f"{k}={v}" for k, v in list(r['values_by_header'].items())[:8])
                for r in records[:50]
            ))
        if pairs:
            workbook_parts.append("[DETECTED KEY/VALUE PAIRS]\n" + "\n".join(pairs[:120]))
        if formula_notes:
            workbook_parts.append("[FORMULA RESULTS]\n" + "\n".join(formula_notes[:80]))

        sheet_summaries.append({
            "sheet_name": sheet_name,
            "max_row_seen": max_row,
            "max_col_seen": max_col,
            "non_empty_rows": sum(1 for row in grid if _is_meaningful_spreadsheet_row(row)),
            "detected_key_value_pairs": len(pairs),
        })

    text = normalize_text("\n\n".join(workbook_parts))
    return {
        "method": "excel_xlsx",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": len(wb_values.sheetnames),
            "sheets": sheet_summaries,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_key_value_pairs": len(all_pairs),
            "detected_record_count": len(all_records),
        },
        "records": all_records,
    }


def extract_xls_text(excel_bytes: bytes, filename: str) -> Dict[str, Any]:
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError("Missing dependency for .xls files. Install: pip install xlrd") from exc

    book = xlrd.open_workbook(file_contents=excel_bytes)

    workbook_parts: List[str] = [
        "[SPREADSHEET TEXT]",
        f"Filename: {filename}",
        f"Workbook sheets: {', '.join(book.sheet_names())}",
        "Instructions for AI: Values are presented with row numbers and cell addresses. Use explicit labels and nearby values.",
    ]
    sheet_summaries: List[Dict[str, Any]] = []
    total_pairs = 0
    all_records: List[Dict[str, Any]] = []

    for sheet in book.sheets():
        max_row = min(sheet.nrows, settings.excel_max_rows_per_sheet)
        max_col = min(sheet.ncols, settings.excel_max_cols_per_sheet)
        grid: List[List[str]] = []
        for r in range(max_row):
            row_vals: List[str] = []
            for c in range(max_col):
                cell = sheet.cell(r, c)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                row_vals.append(normalize_spreadsheet_cell(value))
            grid.append(row_vals)
        records = _extract_table_records_from_grid(sheet.name, grid)
        all_records.extend(records)
        pairs = _detect_key_value_pairs_from_grid(grid, sheet.name)
        total_pairs += len(pairs)
        workbook_parts.append(_grid_to_ai_text(sheet.name, grid, max_col))
        if records:
            workbook_parts.append("[DETECTED SPREADSHEET RECORDS]\n" + "\n".join(
                f"Record {r['record_index']} from {r['sheet_name']} row {r['source_row']}: "
                + " | ".join(f"{k}={v}" for k, v in list(r['values_by_header'].items())[:8])
                for r in records[:50]
            ))
        if pairs:
            workbook_parts.append("[DETECTED KEY/VALUE PAIRS]\n" + "\n".join(pairs[:120]))
        sheet_summaries.append({
            "sheet_name": sheet.name,
            "max_row_seen": max_row,
            "max_col_seen": max_col,
            "non_empty_rows": sum(1 for row in grid if _is_meaningful_spreadsheet_row(row)),
            "detected_key_value_pairs": len(pairs),
        })

    text = normalize_text("\n\n".join(workbook_parts))
    return {
        "method": "excel_xls",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": book.nsheets,
            "sheets": sheet_summaries,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_key_value_pairs": total_pairs,
            "detected_record_count": len(all_records),
        },
        "records": all_records,
    }


def extract_csv_text(csv_bytes: bytes, filename: str) -> Dict[str, Any]:
    text_raw = csv_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text_raw))
    rows = []
    for i, row in enumerate(reader, start=1):
        if i > settings.excel_max_rows_per_sheet:
            break
        clean = [normalize_spreadsheet_cell(v) for v in row[: settings.excel_max_cols_per_sheet]]
        if _is_meaningful_spreadsheet_row(clean):
            rows.append(f"ROW {i}: " + "\t".join(clean))
    text = normalize_text("\n".join([
        "[SPREADSHEET TEXT]",
        f"Filename: {filename}",
        "--- SHEET: CSV ---",
        *rows,
    ]))
    return {
        "method": "csv",
        "text": text,
        "quality": {
            "document_type_detected": "spreadsheet",
            "sheet_count": 1,
            "char_count": len(text),
            "field_hits": count_field_hits(text),
            "detected_record_count": 0,
        },
        "records": [],
    }


def extract_spreadsheet_text_professionally(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    ext = _file_extension(filename)
    if ext == ".xlsx":
        return extract_xlsx_text(file_bytes, filename)
    if ext == ".xls":
        return extract_xls_text(file_bytes, filename)
    if ext == ".csv":
        return extract_csv_text(file_bytes, filename)
    raise ValueError(f"Unsupported spreadsheet type: {ext}")


def extract_document_text_professionally(file_bytes: bytes, filename: str, force_ocr: bool = False) -> Dict[str, Any]:
    from pdf_extractor import extract_pdf_text_professionally

    ext = _file_extension(filename)
    if ext == ".pdf":
        result = extract_pdf_text_professionally(file_bytes, force_ocr=force_ocr)
        result["source_file_type"] = "pdf"
        return result
    if ext in SUPPORTED_SPREADSHEET_EXTENSIONS:
        result = extract_spreadsheet_text_professionally(file_bytes, filename)
        result["source_file_type"] = "spreadsheet"
        return result
    raise ValueError(f"Unsupported file type '{ext}'. Upload PDF, XLSX, XLS, or CSV.")