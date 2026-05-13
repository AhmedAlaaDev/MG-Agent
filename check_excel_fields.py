import openpyxl

filepath = r"C:\Users\AhmedAlaa\Downloads\LCL-HBLs\2311135-MESCO LCL-SZ TO ALEX.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb.active

print("Sheet name:", ws.title)
print()
print("Header row (row 7):")
headers = []
for c in range(1, ws.max_column + 1):
    val = ws.cell(7, c).value
    if val:
        headers.append((c, val))
        print(f"  Col {c}: {val}")

print()
print("Sample data row 8:")
for c, header in headers:
    val = ws.cell(8, c).value
    print(f"  {header}: {val}")