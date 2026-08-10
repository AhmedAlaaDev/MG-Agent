#!/usr/bin/env python3
"""
WE-CAN proxy billing Excel extractor.

Extracts every unique BL/HBL row from WE-CAN/Ameta proxy billing sheets:
- CIF records
- FOB records
- per-record debit/credit processing charges
- CIF/FOB totals
- bottom DEBIT / CREDIT / FINAL DN summary

Usage:
    pip install xlrd openpyxl
    python wecan_proxy_bill_extractor.py "input.xls" -o extracted.json --csv extracted.csv
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_cell(value: Any) -> str:
    """Normalize an Excel cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Several legacy WE-CAN sheets store a few stray seconds in otherwise
        # date-only cells. Treat near-midnight values as dates so ETD does not
        # become e.g. ``2026-05-30 00:00:09``.
        if value.hour == 0 and value.minute == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return (f"{value:.10f}").rstrip("0").rstrip(".")
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return re.sub(r"\s+", " ", text).strip()


def _get(row: List[str], col: int, default: str = "") -> str:
    if col < 0 or col >= len(row):
        return default
    return (row[col] or "").strip()


def _first_non_empty(row: List[str]) -> str:
    return next((v.strip() for v in row if str(v).strip()), "")


def _money_to_float(value: Any) -> Optional[float]:
    """Convert values like '$1,059.80', 'US$3,800.00', 35 to float/None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("US$", "").replace("USD", "").replace("$", "").replace(",", "").strip()
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def _num_to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _is_bl_number(value: str) -> bool:
    """Detect HHSE5073166, QPSE5071283, HHSE5071857YA, etc."""
    value = (value or "").strip().upper()
    return bool(re.fullmatch(r"[A-Z]{2,}[A-Z0-9]*\d{4,}[A-Z0-9]*", value))


def _is_total_row(row: List[str]) -> bool:
    first = _first_non_empty(row).upper()
    return first.startswith("TOTAL") or first.startswith("SUBTOTAL") or first.startswith("GRAND TOTAL")


def _row_text(row: Iterable[str]) -> str:
    return " ".join(str(v).strip() for v in row if str(v).strip())


def _safe_round(value: Optional[float]) -> Optional[float]:
    return None if value is None else round(value, 2)


# ---------------------------------------------------------------------------
# Excel readers
# ---------------------------------------------------------------------------

def read_excel_grids(path: str | Path) -> Dict[str, List[List[str]]]:
    """
    Return {sheet_name: grid}. Supports .xls and .xlsx.

    Dependencies:
      .xls  -> xlrd
      .xlsx -> openpyxl
    """
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise ImportError("For .xls files install: pip install xlrd") from exc

        book = xlrd.open_workbook(str(path))
        result: Dict[str, List[List[str]]] = {}
        for sheet in book.sheets():
            grid: List[List[str]] = []
            for r in range(sheet.nrows):
                row: List[str] = []
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    value: Any = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                        except Exception:
                            pass
                    row.append(normalize_cell(value))
                grid.append(row)
            result[sheet.name] = grid
        return result

    if ext == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise ImportError("For .xlsx files install: pip install openpyxl") from exc

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        result = {}
        for ws in wb.worksheets:
            grid = []
            for row in ws.iter_rows(values_only=True):
                grid.append([normalize_cell(v) for v in row])
            result[ws.title] = grid
        return result

    raise ValueError(f"Unsupported Excel extension: {ext}. Use .xls or .xlsx")


# ---------------------------------------------------------------------------
# WE-CAN proxy-bill column maps
# ---------------------------------------------------------------------------

# Zero-based column positions from the WE-CAN proxy bill layout shown in the file.
CIF_COLS = {
    "hbl": 0,
    "dest": 1,
    "pkgs": 2,
    "gw": 3,
    "cbm": 4,
    "charged_wm": 5,
    "term": 6,
    "agreement_rebate": 7,
    # col 8 is usually a merged/blank helper column
    "pcs": 9,
    "lss": 10,
    "loading_unloading": 11,
    "admin_fees": 12,
    "debit_total": 13,
    "credit_ts": 14,
    "credit_dest_local_prepaid": 15,
    "credit_custom": 16,
    "credit_truckage": 17,
    # Some versions contain extra right-side credit columns.
    # These right-side columns are under "AGREEMENT LOCAL WITH MARINE".
    # They are not invoice credits and are intentionally kept separate from
    # the DEBIT - CREDIT = FINAL DN reconciliation.
    "local_agreement_thc_cfs": 19,
    "local_agreement_do_admin": 20,
    "local_agreement_total": 21,
}

FOB_COLS = {
    "hbl": 0,
    "dest": 1,
    "pkgs": 2,
    "gw": 3,
    "cbm": 4,
    "charged_wm": 5,
    "term": 6,
    "ex_work": 7,
    # col 8 is usually a merged/blank helper column
    "thc": 9,
    # col 10 is usually a merged/blank helper column
    "fob_of": 11,
}


# ---------------------------------------------------------------------------
# Header/meta extraction
# ---------------------------------------------------------------------------

def _find_value_after_label(row: List[str], label_regex: str, max_lookahead: int = 5) -> str:
    pattern = re.compile(label_regex, re.I)
    for i, cell in enumerate(row):
        if pattern.search(cell or ""):
            for value in row[i + 1 : i + 1 + max_lookahead]:
                if str(value).strip():
                    return str(value).strip()
    return ""


def parse_global_header(grid: List[List[str]]) -> Dict[str, Any]:
    meta: Dict[str, Any] = {}

    # Company/logo line
    for row in grid[:5]:
        text = _row_text(row).upper()
        if "WE-CAN" in text or "WE－CAN" in text:
            meta["company"] = "WE-CAN INTERNATIONAL LOGISTICS CO., LTD."
            break

    # Consolidation description, usually a centered line above vessel info.
    for row in grid[:8]:
        text = _row_text(row)
        if "CONSOL BOX" in text.upper():
            meta["consol_description"] = text
            break

    # Search the first rows for labels; this is more robust than fixed row numbers.
    for row in grid[:12]:
        vessel_raw = _find_value_after_label(row, r"\bVSL\b|VESSEL")
        if vessel_raw and not meta.get("vessel"):
            if "/" in vessel_raw:
                vessel, voyage = vessel_raw.split("/", 1)
                meta["vessel"] = vessel.strip()
                meta["voyage"] = voyage.strip()
            else:
                meta["vessel"] = vessel_raw

        etd = _find_value_after_label(row, r"\bETD\b")
        if etd and not meta.get("etd"):
            meta["etd"] = etd

        cntr = _find_value_after_label(row, r"CNTR\.?\s*NO")
        if cntr and not meta.get("container_no"):
            meta["container_no"] = cntr

        carrier = _find_value_after_label(row, r"CARRIER")
        if carrier and not meta.get("carrier"):
            meta["carrier"] = carrier

        mbl = _find_value_after_label(row, r"\bMBL\s*NO\b|MASTER\s*B/?L")
        if mbl and not meta.get("mbl_no"):
            meta["mbl_no"] = mbl

        container_type = _find_value_after_label(row, r"TYPE\s+OF\s+CNTR|CNTR\s+TYPE")
        if container_type and not meta.get("container_type"):
            meta["container_type"] = container_type

        total_gw = _find_value_after_label(row, r"TOTAL\s+G\.?W")
        if total_gw and not meta.get("total_gw_kgs"):
            meta["total_gw_kgs"] = _num_to_float(total_gw)

        total_volume = _find_value_after_label(row, r"TOTAL\s+VOLUME")
        if total_volume and not meta.get("total_volume_cbm"):
            meta["total_volume_cbm"] = _num_to_float(total_volume)

        ocean_rate = _find_value_after_label(row, r"OCEAN\s+FREIGHT")
        if ocean_rate and not meta.get("ocean_freight"):
            meta["ocean_freight"] = _money_to_float(ocean_rate)

    return {k: v for k, v in meta.items() if v not in (None, "")}


# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------

def build_mesco_payload(record: Dict[str, Any], meta: Dict[str, Any]) -> Dict[str, Any]:
    hbl = record.get("hbl_no")
    return {
        "document_type": "Bill of Lading",
        "mesco_masterblno": meta.get("mbl_no"),
        "mesco_houseblno": hbl,
        "mesco_shippernamecontactno": meta.get("company"),
        "mesco_vessel": meta.get("vessel"),
        "mesco_voytruckno": meta.get("voyage"),
        "mesco_destination": record.get("dest") or "Alexandria",
        "mesco_cargodescription": meta.get("consol_description"),
        "cr401_totalgrossweight": meta.get("total_gw_kgs"),
        "cr401_totalvolume": meta.get("total_volume_cbm"),
        "cr401_totalpackages": record.get("pkgs"),
        "mesco_containertype": meta.get("container_type"),
        "mesco_pcfreightterm": record.get("term"),
        "mesco_etdorigin": meta.get("etd"),
        "mesco_incoterm": record.get("term"),
        "container_number": meta.get("container_no"),
        "containers": [
            {
                "container_number": meta.get("container_no"),
                "container_type": meta.get("container_type"),
                "packages": record.get("pkgs"),
                "gross_weight_kg": record.get("gw_kgs"),
                "measurement_cbm": record.get("volume_cbm"),
            }
        ],
    }


def extract_cif_record(row: List[str], source_row: int, sheet_name: str, meta: Dict[str, Any]) -> Dict[str, Any]:
    record = {
        "sheet_name": sheet_name,
        "source_row": source_row,
        "cargo_type": "CIF",
        "hbl_no": _get(row, CIF_COLS["hbl"]).upper(),
        "dest": _get(row, CIF_COLS["dest"]),
        "pkgs": _num_to_float(_get(row, CIF_COLS["pkgs"])),
        "gw_kgs": _num_to_float(_get(row, CIF_COLS["gw"])),
        "volume_cbm": _num_to_float(_get(row, CIF_COLS["cbm"])),
        "charged_wm": _num_to_float(_get(row, CIF_COLS["charged_wm"])),
        "term": _get(row, CIF_COLS["term"]),
    }

    debit = {
        "agreement_rebate": _money_to_float(_get(row, CIF_COLS["agreement_rebate"])),
        "pcs": _money_to_float(_get(row, CIF_COLS["pcs"])),
        "lss": _money_to_float(_get(row, CIF_COLS["lss"])),
        "loading_unloading": _money_to_float(_get(row, CIF_COLS["loading_unloading"])),
        "admin_fees": _money_to_float(_get(row, CIF_COLS["admin_fees"])),
        "total": _money_to_float(_get(row, CIF_COLS["debit_total"])),
    }
    credit = {
        "ts": _money_to_float(_get(row, CIF_COLS["credit_ts"])),
        "dest_local_prepaid": _money_to_float(_get(row, CIF_COLS["credit_dest_local_prepaid"])),
        "custom": _money_to_float(_get(row, CIF_COLS["credit_custom"])),
        "truckage": _money_to_float(_get(row, CIF_COLS["credit_truckage"])),
    }
    local_agreement = {
        "thc_cfs": _money_to_float(_get(row, CIF_COLS["local_agreement_thc_cfs"])),
        "do_admin": _money_to_float(_get(row, CIF_COLS["local_agreement_do_admin"])),
        "total": _money_to_float(_get(row, CIF_COLS["local_agreement_total"])),
    }

    # Remove empty money fields, but preserve known zeros if present in the sheet.
    debit = {k: v for k, v in debit.items() if v is not None}
    credit = {k: v for k, v in credit.items() if v is not None}
    local_agreement = {k: v for k, v in local_agreement.items() if v is not None}

    return {
        "unique_key": record["hbl_no"],
        "record": record,
        "financial_processing": {
            **record,
            "debit": debit,
            "credit": credit,
            "local_agreement": local_agreement,
        },
        "mesco_payload": build_mesco_payload(record, meta),
    }


def extract_fob_record(row: List[str], source_row: int, sheet_name: str, meta: Dict[str, Any]) -> Dict[str, Any]:
    record = {
        "sheet_name": sheet_name,
        "source_row": source_row,
        "cargo_type": "FOB",
        "hbl_no": _get(row, FOB_COLS["hbl"]).upper(),
        "dest": _get(row, FOB_COLS["dest"]),
        "pkgs": _num_to_float(_get(row, FOB_COLS["pkgs"])),
        "gw_kgs": _num_to_float(_get(row, FOB_COLS["gw"])),
        "volume_cbm": _num_to_float(_get(row, FOB_COLS["cbm"])),
        "charged_wm": _num_to_float(_get(row, FOB_COLS["charged_wm"])),
        "term": _get(row, FOB_COLS["term"]),
    }

    debit = {
        "ex_work": _money_to_float(_get(row, FOB_COLS["ex_work"])),
        "thc": _money_to_float(_get(row, FOB_COLS["thc"])),
        "fob_of": _money_to_float(_get(row, FOB_COLS["fob_of"])),
    }
    debit = {k: v for k, v in debit.items() if v is not None}

    return {
        "unique_key": record["hbl_no"],
        "record": record,
        "financial_processing": {**record, "debit": debit, "credit": {}},
        "mesco_payload": build_mesco_payload(record, meta),
    }


def parse_section_records(
    grid: List[List[str]],
    sheet_name: str,
    meta: Dict[str, Any],
    cargo_type: str,
    start_row_0: int,
    end_row_0: int,
) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for ri in range(start_row_0, min(end_row_0, len(grid))):
        row = grid[ri]
        if not any(str(v).strip() for v in row):
            continue
        if _is_total_row(row):
            continue
        hbl = _get(row, 0)
        if not _is_bl_number(hbl):
            continue
        if cargo_type == "CIF":
            records.append(extract_cif_record(row, ri + 1, sheet_name, meta))
        else:
            records.append(extract_fob_record(row, ri + 1, sheet_name, meta))
    return records


# ---------------------------------------------------------------------------
# Totals and summary extraction
# ---------------------------------------------------------------------------

def parse_totals_and_final_summary(grid: List[List[str]]) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "cif_total": {},
        "fob_total": {},
        "final_summary": {},
    }

    for ri, row in enumerate(grid):
        text = _row_text(row).upper()
        non_empty = [str(v).strip() for v in row if str(v).strip()]

        if "TOTAL CIF CARGO" in text:
            result["cif_total"] = {
                "source_row": ri + 1,
                "pkgs": _num_to_float(_get(row, CIF_COLS["pkgs"])),
                "gw_kgs": _num_to_float(_get(row, CIF_COLS["gw"])),
                "volume_cbm": _num_to_float(_get(row, CIF_COLS["cbm"])),
                "charged_wm": _num_to_float(_get(row, CIF_COLS["charged_wm"])),
                "debit_total": _money_to_float(_get(row, CIF_COLS["debit_total"])),
                "credit_ts": _money_to_float(_get(row, CIF_COLS["credit_ts"])),
                "credit_dest_local_prepaid": _money_to_float(_get(row, CIF_COLS["credit_dest_local_prepaid"])),
                "credit_custom": _money_to_float(_get(row, CIF_COLS["credit_custom"])),
                "credit_truckage": _money_to_float(_get(row, CIF_COLS["credit_truckage"])),
            }
            result["cif_total"] = {k: v for k, v in result["cif_total"].items() if v is not None or k == "source_row"}

        elif "TOTAL FOB CARGO" in text:
            result["fob_total"] = {
                "source_row": ri + 1,
                "pkgs": _num_to_float(_get(row, FOB_COLS["pkgs"])),
                "gw_kgs": _num_to_float(_get(row, FOB_COLS["gw"])),
                "volume_cbm": _num_to_float(_get(row, FOB_COLS["cbm"])),
                "charged_wm": _num_to_float(_get(row, FOB_COLS["charged_wm"])),
                "ex_work_total": _money_to_float(_get(row, FOB_COLS["ex_work"])),
                "thc_total": _money_to_float(_get(row, FOB_COLS["thc"])),
                "fob_of_total": _money_to_float(_get(row, FOB_COLS["fob_of"])),
            }
            result["fob_total"] = {k: v for k, v in result["fob_total"].items() if v is not None or k == "source_row"}

        if not non_empty:
            continue

        label = non_empty[0].upper()
        amount = None
        for cell in reversed(non_empty):
            amount = _money_to_float(cell)
            if amount is not None:
                break

        if label == "DEBIT" and amount is not None:
            result["final_summary"]["debit"] = amount
            result["final_summary"]["debit_source_row"] = ri + 1
        elif label == "CREDIT" and amount is not None:
            result["final_summary"]["credit"] = amount
            result["final_summary"]["credit_source_row"] = ri + 1
        elif "FINAL DN" in label and amount is not None:
            result["final_summary"]["final_dn"] = amount
            result["final_summary"]["final_dn_label"] = non_empty[0]
            result["final_summary"]["final_dn_source_row"] = ri + 1

    debit = result["final_summary"].get("debit")
    credit = result["final_summary"].get("credit")
    if debit is not None and credit is not None:
        result["final_summary"]["calculated_final_dn"] = round(debit - credit, 2)

    return result


def calculate_totals_from_records(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Independent totals calculated from unique extracted records."""
    calc = {
        "record_count": len(records),
        "cif_count": 0,
        "fob_count": 0,
        "cif_debit_total": 0.0,
        "cif_credit_total": 0.0,
        "fob_ex_work_total": 0.0,
        "fob_thc_total": 0.0,
        "fob_of_total": 0.0,
        "grand_debit_total": 0.0,
        "grand_credit_total": 0.0,
        "local_agreement_total": 0.0,
        "calculated_final_dn": 0.0,
    }

    for item in records:
        fp = item["financial_processing"]
        cargo_type = fp["cargo_type"]
        debit = fp.get("debit", {})
        credit = fp.get("credit", {})
        local_agreement = fp.get("local_agreement", {})

        if cargo_type == "CIF":
            calc["cif_count"] += 1
            cif_debit = debit.get("total") or 0.0
            cif_credit = sum(v for v in credit.values() if isinstance(v, (int, float)))
            calc["cif_debit_total"] += cif_debit
            calc["cif_credit_total"] += cif_credit or 0.0
            calc["local_agreement_total"] += local_agreement.get("total") or 0.0

        elif cargo_type == "FOB":
            calc["fob_count"] += 1
            calc["fob_ex_work_total"] += debit.get("ex_work") or 0.0
            calc["fob_thc_total"] += debit.get("thc") or 0.0
            calc["fob_of_total"] += debit.get("fob_of") or 0.0

    calc["grand_debit_total"] = calc["cif_debit_total"] + calc["fob_ex_work_total"] + calc["fob_thc_total"] + calc["fob_of_total"]
    calc["grand_credit_total"] = calc["cif_credit_total"]
    calc["calculated_final_dn"] = calc["grand_debit_total"] - calc["grand_credit_total"]

    for key, value in list(calc.items()):
        if isinstance(value, float):
            calc[key] = round(value, 2)

    return calc


# ---------------------------------------------------------------------------
# Invoice-agent / Dynamics adapter
# ---------------------------------------------------------------------------

_DEBIT_TERM_LABELS = {
    "agreement_rebate": "AGREEMENT REBATE",
    "pcs": "PCS",
    "lss": "LSS",
    "loading_unloading": "LOADING/UNLOADING",
    "admin_fees": "ADMIN FEES",
    "ex_work": "EX-WORK",
    "thc": "THC",
    "fob_of": "FOB OCEAN FREIGHT",
}

_CREDIT_TERM_LABELS = {
    "ts": "CREDIT - T/S",
    "dest_local_prepaid": "CREDIT - DESTINATION LOCAL PREPAID",
    "custom": "CREDIT - DAP/DDU/DDP CUSTOMS",
    "truckage": "CREDIT - DAP/DDU/DDP TRUCKAGE",
}

_LOCAL_AGREEMENT_LABELS = {
    "thc_cfs": "LOCAL AGREEMENT - THC/CFS",
    "do_admin": "LOCAL AGREEMENT - D/O + ADMIN",
}


def _invoice_reference_from_filename(filename: str) -> str:
    stem = Path(filename or "invoice").stem
    for pattern in (r"_Get_(\d+)", r"#(\d+)"):
        match = re.search(pattern, stem, re.I)
        if match:
            return match.group(1)
    return stem[:100]


def _rated_line_item(
    description: str,
    amount: float,
    record: Dict[str, Any],
    *,
    direction: str,
    term_key: str,
) -> Dict[str, Any]:
    charged_wm = float(record.get("charged_wm") or 0)
    quantity = 1.0
    unit_price = float(amount)

    # Preserve the explicit W/M rate where the sheet's amount proves it. This
    # gives Dynamics meaningful quantity and unit-rate fields while retaining
    # the exact spreadsheet total.
    known_rate = {
        "agreement_rebate": 75.0,
        "pcs": 15.0,
        "loading_unloading": 65.0,
        "fob_of": 79.28202200868931,
        "ts": 20.0,
        "thc_cfs": 30.0,
    }.get(term_key)
    sign = -1.0 if direction == "credit" else 1.0
    if known_rate and charged_wm and abs(abs(amount) - charged_wm * known_rate) <= 0.02:
        quantity = charged_wm
        unit_price = sign * known_rate
    else:
        unit_price = sign * abs(float(amount))

    return {
        "service_description": description,
        "quantity": quantity,
        "unit_price": round(unit_price, 8),
        "currency": "USD",
        "total_amount": round(sign * abs(float(amount)), 2),
        "posting_direction": direction,
        "source_term": term_key,
        "source_row": record.get("source_row"),
    }


def _record_to_invoice_group(item: Dict[str, Any], meta: Dict[str, Any]) -> Dict[str, Any]:
    record = item["record"]
    processing = item["financial_processing"]
    debit = processing.get("debit", {})
    credit = processing.get("credit", {})
    local_agreement = processing.get("local_agreement", {})

    invoice_lines: List[Dict[str, Any]] = []
    for key, description in _DEBIT_TERM_LABELS.items():
        amount = debit.get(key)
        if amount is not None and abs(float(amount)) > 0:
            invoice_lines.append(
                _rated_line_item(description, float(amount), record, direction="debit", term_key=key)
            )
    for key, description in _CREDIT_TERM_LABELS.items():
        amount = credit.get(key)
        if amount is not None and abs(float(amount)) > 0:
            invoice_lines.append(
                _rated_line_item(description, float(amount), record, direction="credit", term_key=key)
            )

    local_lines: List[Dict[str, Any]] = []
    for key, description in _LOCAL_AGREEMENT_LABELS.items():
        amount = local_agreement.get(key)
        if amount is not None and abs(float(amount)) > 0:
            local_lines.append(
                _rated_line_item(description, float(amount), record, direction="local_agreement", term_key=key)
            )

    debit_total = sum(
        float(v) for key, v in debit.items()
        if key != "total" and isinstance(v, (int, float))
    )
    credit_total = sum(float(v) for v in credit.values() if isinstance(v, (int, float)))
    net_total = round(debit_total - credit_total, 2)

    return {
        "house_bl_number": record.get("hbl_no"),
        "currency": "USD",
        "shipment_ref": record.get("hbl_no"),
        "container_number": meta.get("container_no"),
        "container_type": meta.get("container_type"),
        "cbm": record.get("volume_cbm"),
        "kgs": record.get("gw_kgs"),
        "packages": record.get("pkgs"),
        "charged_wm": record.get("charged_wm"),
        "term": record.get("term"),
        "destination": record.get("dest"),
        "source_row": record.get("source_row"),
        "debit_total": round(debit_total, 2),
        "credit_total": round(credit_total, 2),
        "total_amount": net_total,
        "line_items": invoice_lines,
        # Extracted and returned for audit, but deliberately not included in
        # invoice line_items: this block is outside DEBIT/CREDIT and is not
        # part of FINAL DN in the source workbook.
        "local_agreement_items": local_lines,
        "local_agreement_total": round(
            float(local_agreement.get("total") or sum(
                float(v) for key, v in local_agreement.items()
                if key != "total" and isinstance(v, (int, float))
            )),
            2,
        ),
    }


def to_multi_invoice_payload(result: Dict[str, Any], source_filename: str) -> Dict[str, Any]:
    """Map a parsed WE-CAN proxy workbook to the multi-invoice API schema."""
    sheets = result.get("sheets") or []
    if not sheets or not result.get("records"):
        raise ValueError("The Excel file is not a recognized WE-CAN proxy billing workbook.")

    primary_sheet = sheets[0]
    meta = primary_sheet.get("meta") or {}
    groups = [_record_to_invoice_group(item, meta) for item in result["records"]]
    summary = primary_sheet.get("processing_summary") or {}
    calculated = primary_sheet.get("calculated_totals_from_records") or {}
    final_summary = summary.get("final_summary") or {}

    posted_net = round(sum(
        float(line.get("total_amount") or 0)
        for group in groups
        for line in group.get("line_items") or []
    ), 2)
    stated_final = final_summary.get("final_dn")
    validation = {
        "record_count": len(groups),
        "line_items_count": sum(len(group["line_items"]) for group in groups),
        "calculated_debit": calculated.get("grand_debit_total"),
        "calculated_credit": calculated.get("grand_credit_total"),
        "calculated_final_dn": calculated.get("calculated_final_dn"),
        "stated_final_dn": stated_final,
        "posting_net_total": posted_net,
        "local_agreement_total": calculated.get("local_agreement_total"),
        "is_reconciled": stated_final is not None and abs(float(stated_final) - posted_net) <= 0.02,
    }
    if not validation["is_reconciled"]:
        raise ValueError(
            "WE-CAN invoice totals did not reconcile: "
            f"posting lines={posted_net}, stated FINAL DN={stated_final}."
        )

    return {
        "document_type": "WE-CAN Proxy Billing Excel",
        "vendor_name": meta.get("company") or "WE-CAN INTERNATIONAL LOGISTICS CO., LTD.",
        "vendor_invoice_number": _invoice_reference_from_filename(source_filename),
        "master_bl_number": meta.get("mbl_no"),
        "container_number": meta.get("container_no"),
        "container_type": meta.get("container_type"),
        "currency": "USD",
        "vessel_name": meta.get("vessel"),
        "voyage_number": meta.get("voyage"),
        "etd": meta.get("etd"),
        "carrier": meta.get("carrier"),
        "ocean_freight": meta.get("ocean_freight"),
        "total_gross_weight_kg": meta.get("total_gw_kgs"),
        "total_volume_cbm": meta.get("total_volume_cbm"),
        "groups": groups,
        "processing_summary": summary,
        "extraction_validation": validation,
    }


# ---------------------------------------------------------------------------
# Main extraction logic
# ---------------------------------------------------------------------------

def find_proxy_sections(grid: List[List[str]]) -> Tuple[Optional[int], Optional[int]]:
    cif_section_row = None
    fob_section_row = None

    for i, row in enumerate(grid):
        text = _row_text(row).upper()
        if "CIF CARGO" in text and cif_section_row is None:
            cif_section_row = i
        if "FOB CARGO" in text and fob_section_row is None:
            fob_section_row = i

    return cif_section_row, fob_section_row


def dedupe_records_by_bl(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    unique: List[Dict[str, Any]] = []
    duplicates: List[Dict[str, Any]] = []
    seen = set()

    for rec in records:
        hbl = rec.get("unique_key") or rec.get("record", {}).get("hbl_no")
        key = str(hbl or "").strip().upper()
        if not key:
            continue
        if key in seen:
            duplicates.append(rec)
            continue
        seen.add(key)
        unique.append(rec)

    # Re-number after de-duplication.
    for idx, rec in enumerate(unique, start=1):
        rec["record_index"] = idx

    return unique, duplicates


def extract_wecan_proxy_bill_from_grid(grid: List[List[str]], sheet_name: str = "Sheet1") -> Dict[str, Any]:
    meta = parse_global_header(grid)
    cif_section_row, fob_section_row = find_proxy_sections(grid)

    all_records: List[Dict[str, Any]] = []

    if cif_section_row is not None:
        # CIF layout:
        # row 0: CIF CARGO'S INFORMATION / DEBIT / CREDIT
        # row 1: major headers
        # row 2: subheaders
        # row 3: first data row
        cif_data_start = cif_section_row + 4
        cif_data_end = fob_section_row if fob_section_row is not None else len(grid)
        all_records.extend(parse_section_records(grid, sheet_name, meta, "CIF", cif_data_start, cif_data_end))

    if fob_section_row is not None:
        # FOB layout:
        # row 0: FOB CARGO'S INFORMATION / DEBIT
        # row 1: headers
        # row 2: first data row
        fob_data_start = fob_section_row + 2
        fob_data_end = len(grid)
        all_records.extend(parse_section_records(grid, sheet_name, meta, "FOB", fob_data_start, fob_data_end))

    unique_records, duplicate_records = dedupe_records_by_bl(all_records)
    totals_from_sheet = parse_totals_and_final_summary(grid)
    totals_from_records = calculate_totals_from_records(unique_records)

    return {
        "sheet_name": sheet_name,
        "meta": meta,
        "record_count": len(unique_records),
        "unique_bl_numbers": [rec["record"]["hbl_no"] for rec in unique_records],
        "records": unique_records,
        "duplicate_records_skipped": [rec.get("record", {}).get("hbl_no") for rec in duplicate_records],
        "processing_summary": totals_from_sheet,
        "calculated_totals_from_records": totals_from_records,
    }


def extract_wecan_proxy_bill(path: str | Path) -> Dict[str, Any]:
    grids = read_excel_grids(path)
    sheets: List[Dict[str, Any]] = []

    for sheet_name, grid in grids.items():
        cif_row, fob_row = find_proxy_sections(grid)
        # Only parse sheets that look like the WE-CAN proxy bill.
        looks_like_proxy = cif_row is not None or fob_row is not None or any(
            "WE-CAN" in _row_text(row).upper() and "LOGISTICS" in _row_text(row).upper()
            for row in grid[:15]
        )
        if not looks_like_proxy:
            continue
        sheets.append(extract_wecan_proxy_bill_from_grid(grid, sheet_name))

    # Flatten records across sheets and guarantee global uniqueness too.
    flat_records: List[Dict[str, Any]] = []
    for sheet in sheets:
        flat_records.extend(sheet["records"])
    flat_records, duplicate_records = dedupe_records_by_bl(flat_records)

    return {
        "source_file": str(path),
        "sheet_count_parsed": len(sheets),
        "record_count": len(flat_records),
        "unique_bl_numbers": [rec["record"]["hbl_no"] for rec in flat_records],
        "records": flat_records,
        "sheets": sheets,
        "duplicate_records_skipped_global": [rec.get("record", {}).get("hbl_no") for rec in duplicate_records],
    }


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_records_csv(result: Dict[str, Any], csv_path: str | Path) -> None:
    rows: List[Dict[str, Any]] = []
    for rec in result.get("records", []):
        fp = rec["financial_processing"]
        debit = fp.get("debit", {})
        credit = fp.get("credit", {})
        local_agreement = fp.get("local_agreement", {})
        base = {
            "record_index": rec.get("record_index"),
            "sheet_name": fp.get("sheet_name"),
            "source_row": fp.get("source_row"),
            "cargo_type": fp.get("cargo_type"),
            "hbl_no": fp.get("hbl_no"),
            "dest": fp.get("dest"),
            "pkgs": fp.get("pkgs"),
            "gw_kgs": fp.get("gw_kgs"),
            "volume_cbm": fp.get("volume_cbm"),
            "charged_wm": fp.get("charged_wm"),
            "term": fp.get("term"),
        }
        for k, v in debit.items():
            base[f"debit_{k}"] = v
        for k, v in credit.items():
            base[f"credit_{k}"] = v
        for k, v in local_agreement.items():
            base[f"local_agreement_{k}"] = v
        rows.append(base)

    # Stable preferred column order, plus any discovered columns.
    preferred = [
        "record_index", "sheet_name", "source_row", "cargo_type", "hbl_no", "dest",
        "pkgs", "gw_kgs", "volume_cbm", "charged_wm", "term",
        "debit_agreement_rebate", "debit_pcs", "debit_lss", "debit_loading_unloading",
        "debit_admin_fees", "debit_total", "debit_ex_work", "debit_thc", "debit_fob_of",
        "credit_ts", "credit_dest_local_prepaid", "credit_custom", "credit_truckage",
        "local_agreement_thc_cfs", "local_agreement_do_admin", "local_agreement_total",
    ]
    all_keys = []
    for row in rows:
        for key in row:
            if key not in all_keys:
                all_keys.append(key)
    fieldnames = [c for c in preferred if c in all_keys] + [c for c in all_keys if c not in preferred]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Extract unique BL records from WE-CAN proxy billing Excel files.")
    parser.add_argument("input_file", help="Path to .xls or .xlsx file")
    parser.add_argument("-o", "--output-json", default="wecan_proxy_bill_extracted.json", help="Output JSON path")
    parser.add_argument("--csv", dest="output_csv", default=None, help="Optional output CSV path")
    args = parser.parse_args()

    result = extract_wecan_proxy_bill(args.input_file)

    with open(args.output_json, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    if args.output_csv:
        write_records_csv(result, args.output_csv)

    print(json.dumps({
        "record_count": result["record_count"],
        "unique_bl_numbers": result["unique_bl_numbers"],
        "json": args.output_json,
        "csv": args.output_csv,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
